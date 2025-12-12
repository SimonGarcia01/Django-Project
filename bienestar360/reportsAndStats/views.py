from django.views.generic import TemplateView, View
from django.db.models import Count, Avg, Q, F, Max, Min
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.views.generic import TemplateView
from django.utils import timezone
from datetime import datetime, date, timedelta
from login.models import CustomUser, Faculty
from activities.models import Activity, Enrollment, Schedule, ActivityReview, Participation
from django.contrib.auth.models import Group
import csv
import json


def get_admin_user_ids():
    """
    Helper function to get all admin/CADI user IDs that should be excluded from reports.
    Admin users are identified by: admin group, is_staff=True, is_superuser=True, or CADI faculty.
    """
    admin_user_ids = set()
    
    # Get users in admin group
    admin_group = Group.objects.filter(name='admin').first()
    if admin_group:
        admin_user_ids.update(admin_group.user_set.values_list('id', flat=True))
    
    # Get users with staff/superuser flags or CADI faculty
    admin_user_ids.update(
        CustomUser.objects.filter(
            Q(is_staff=True) | Q(is_superuser=True) | Q(faculty__name='CADI')
        ).values_list('id', flat=True)
    )
    
    return admin_user_ids


def get_semester_dates(semester_str):
    """
    Convierte un string de semestre (ej: '2025-01' o '2025-02') en fechas de inicio y fin.
    Formato: year-01 (primer semestre: enero-junio) o year-02 (segundo semestre: julio-diciembre)
    """
    try:
        year, semester = semester_str.split('-')
        year = int(year)
        semester = int(semester)
        
        if semester == 1:
            # Primer semestre: Enero (1) a Junio (6)
            start_date = date(year, 1, 1)
            end_date = date(year, 6, 30)
        elif semester == 2:
            # Segundo semestre: Julio (7) a Diciembre (12)
            start_date = date(year, 7, 1)
            end_date = date(year, 12, 31)
        else:
            return None, None
        
        return start_date, end_date
    except (ValueError, AttributeError):
        return None, None


def get_year_dates(year_str):
    """Convierte un año (ej: '2025') en fechas de inicio y fin del año."""
    try:
        year = int(year_str)
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        return start_date, end_date
    except (ValueError, AttributeError):
        return None, None


def get_semester_from_date(d):
    """Obtiene el semestre (formato 'year-01' o 'year-02') de una fecha."""
    if not d:
        return None
    if isinstance(d, datetime):
        d = d.date()
    year = d.year
    month = d.month
    if 1 <= month <= 6:
        return f"{year}-01"
    elif 7 <= month <= 12:
        return f"{year}-02"
    return None


class GeneralReportsView(TemplateView):
    """
    Diagnostic/general reports view — returns real counts and a few sample rows
    to help identify why earlier counts were zero.
    """
    template_name = 'reportsAndStats/generalReportsAndStatsView.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Exclude admin users from general stats
        admin_user_ids = get_admin_user_ids()
        
        # Basic counts (exclude admin users where applicable)
        total_users = CustomUser.objects.exclude(id__in=admin_user_ids).count()
        total_faculties = Faculty.objects.count()
        total_activities = Activity.objects.count()
        total_enrollments = Enrollment.objects.exclude(user__id__in=admin_user_ids).count()
        total_schedules = Schedule.objects.count()
        total_reviews = ActivityReview.objects.count()

        published_activities = Activity.objects.filter(is_published=True).count() if hasattr(Activity, "is_published") else 0
        requires_registration = Activity.objects.filter(requires_registration=True).count() if hasattr(Activity, "requires_registration") else 0

        activities_with_participants = (
            Activity.objects.annotate(num_participants=Count('enrollments')).filter(num_participants__gt=0).count()
        )

        avg_enrollments_per_activity = round(total_enrollments / total_activities, 2) if total_activities > 0 else 0
        avg_rating = ActivityReview.objects.aggregate(avg=Avg('rating'))['avg'] or 0
        unread_reviews = ActivityReview.objects.filter(is_read=False).count()

        users_sample = list(CustomUser.objects.exclude(id__in=admin_user_ids).values('id', 'username', 'first_name', 'last_name')[:8])
        faculties_sample = list(Faculty.objects.values('id', 'name')[:8])
        activities_sample = list(Activity.objects.values('activityId', 'name', 'type', 'category')[:8])

        top_activities = (
            Activity.objects.annotate(num_participants=Count('enrollments'))
            .order_by('-num_participants')[:10]
        )

        faculty_counts = []
        for f in Faculty.objects.all():
            user_count = f.users.count()
            faculty_enrollments = Enrollment.objects.filter(user__faculty=f).exclude(user__id__in=admin_user_ids).count()
            faculty_counts.append({'name': f.name, 'users': user_count, 'enrollments': faculty_enrollments})

        faculty_distribution = {
            f.name: f.users.count() for f in Faculty.objects.all()
        }

        top_users = (
            CustomUser.objects.exclude(id__in=admin_user_ids).annotate(num_enrollments=Count('enrollments'))
            .order_by('-num_enrollments')[:5]
        )

        context.update({
            'total_users': total_users,
            'total_faculties': total_faculties,
            'total_activities': total_activities,
            'total_enrollments': total_enrollments,
            'total_schedules': total_schedules,
            'total_reviews': total_reviews,
            'published_activities': published_activities,
            'requires_registration': requires_registration,
            'activities_with_participants': activities_with_participants,
            'avg_enrollments_per_activity': avg_enrollments_per_activity,
            'avg_rating': round(avg_rating, 2),
            'unread_reviews': unread_reviews,
            'users_sample': users_sample,
            'faculties_sample': faculties_sample,
            'activities_sample': activities_sample,
            'top_activities': top_activities,
            'faculty_distribution': faculty_distribution,
            'top_users': top_users,
        })

        return context


# =============================================================
# Vista para filtrar y comparar métricas reales
# =============================================================

def download_table_excel(request):
    """
    Descarga la tabla de reportes filtrados en formato Excel.
    """
    try:
        selected_filter = request.GET.get('filter', 'actividad')
        
        # Validar que el filtro sea válido
        valid_filters = ['actividad', 'facultad', 'genero']
        if selected_filter not in valid_filters:
            return JsonResponse({'error': 'No fue posible generar el archivo, por favor intente nuevamente'}, status=400)
        
        data = {}
        
        if selected_filter == 'actividad':
            for act in Activity.objects.all():
                enrolled_users = act.enrollments.all()
                data[act.name] = {
                    'participacion': enrolled_users.count(),
                    'actividades': enrolled_users.count()
                }
        elif selected_filter == 'facultad':
            faculties = Faculty.objects.all()
            for fac in faculties:
                users_in_faculty = CustomUser.objects.filter(faculty=fac)
                total_activities = Enrollment.objects.filter(user__in=users_in_faculty).count()
                data[fac.name] = {
                    'participacion': users_in_faculty.count(),
                    'actividades': total_activities
                }
        elif selected_filter == 'genero':
            genders = CustomUser.objects.values_list('gender', flat=True).distinct()
            for g in genders:
                if g:  # Solo incluir géneros no nulos
                    users_with_gender = CustomUser.objects.filter(gender=g)
                    total_activities = Enrollment.objects.filter(user__in=users_with_gender).count()
                    data[g] = {
                        'participacion': users_with_gender.count(),
                        'actividades': total_activities
                    }
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Participación"
        
        # Estilos para encabezados
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        headers = [selected_filter.title(), 'Participación', 'Actividades']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row_num, (key, info) in enumerate(data.items(), 2):
            ws.cell(row=row_num, column=1, value=str(key) if key else 'N/A')
            ws.cell(row=row_num, column=2, value=info['participacion'])
            ws.cell(row=row_num, column=3, value=info['actividades'])
        
        if not data:
            ws.cell(row=2, column=1, value='No hay datos disponibles')
            ws.cell(row=2, column=2, value=0)
            ws.cell(row=2, column=3, value=0)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reporte_participacion_{selected_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating Excel file: {str(e)}")
        return JsonResponse({'error': 'No fue posible generar el archivo, por favor intente nuevamente'}, status=500)


def download_table_csv(request):
    """
    Descarga la tabla de reportes filtrados en formato CSV.
    """
    try:
        selected_filter = request.GET.get('filter', 'actividad')
        
        # Validar que el filtro sea válido
        valid_filters = ['actividad', 'facultad', 'genero']
        if selected_filter not in valid_filters:
            return JsonResponse({'error': 'No fue posible generar el archivo, por favor intente nuevamente'}, status=400)
        
        data = {}
        
        if selected_filter == 'actividad':
            for act in Activity.objects.all():
                enrolled_users = act.enrollments.all()
                data[act.name] = {
                    'participacion': enrolled_users.count(),
                    'actividades': enrolled_users.count()
                }
        elif selected_filter == 'facultad':
            faculties = Faculty.objects.all()
            for fac in faculties:
                users_in_faculty = CustomUser.objects.filter(faculty=fac)
                total_activities = Enrollment.objects.filter(user__in=users_in_faculty).count()
                data[fac.name] = {
                    'participacion': users_in_faculty.count(),
                    'actividades': total_activities
                }
        elif selected_filter == 'genero':
            genders = CustomUser.objects.values_list('gender', flat=True).distinct()
            for g in genders:
                if g:  # Solo incluir géneros no nulos
                    users_with_gender = CustomUser.objects.filter(gender=g)
                    total_activities = Enrollment.objects.filter(user__in=users_with_gender).count()
                    data[g] = {
                        'participacion': users_with_gender.count(),
                        'actividades': total_activities
                    }
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"reporte_participacion_{selected_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([selected_filter.title(), 'Participación', 'Actividades'])
        
        if data:
            for key, info in data.items():
                writer.writerow([str(key) if key else 'N/A', info['participacion'], info['actividades']])
        else:
            # Si no hay datos, agregar una fila indicando que no hay datos
            writer.writerow(['No hay datos disponibles', 0, 0])
        
        return response
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating CSV file: {str(e)}")
        return JsonResponse({'error': 'No fue posible generar el archivo, por favor intente nuevamente'}, status=500)


class FilteredReportsView(TemplateView):
    """
    Vista para filtrar y mostrar reportes de participación de usuarios
    en actividades y por otros filtros como facultad o género.
    Incluye estadísticas detalladas para cada filtro.
    """
    template_name = 'reportsAndStats/filteredReports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener filtro seleccionado desde GET
        selected_filter = self.request.GET.get('filter', 'actividad')  # default: 'actividad'
        
        # Diccionario que contendrá los datos a mostrar
        data = {}
        
        # Mapeo de géneros para mostrar nombres legibles
        GENDER_NAMES = {
            'M': 'Masculino',
            'F': 'Femenino',
            'O': 'Otro',
        }
        
        # Exclude admin/CADI users from all reports
        admin_user_ids = get_admin_user_ids()
        
        # 🔹 Filtro por actividad
        if selected_filter == 'actividad':
            for act in Activity.objects.all().prefetch_related('enrollments__user', 'enrollments__user__faculty', 
                                                               'reviews', 'schedules'):
                enrollments = act.enrollments.exclude(user__id__in=admin_user_ids)
                participants = enrollments.values('user').distinct().count()
                participations = Participation.objects.filter(activity=act).exclude(user__id__in=admin_user_ids).count()
                
                # Gender breakdown
                gender_breakdown = {}
                for gender_code in ['M', 'F', 'O']:
                    gender_count = enrollments.filter(user__gender=gender_code).values('user').distinct().count()
                    if gender_count > 0:
                        gender_breakdown[GENDER_NAMES.get(gender_code, gender_code)] = gender_count
                
                # Faculty breakdown
                faculty_breakdown = {}
                faculty_enrollments = enrollments.values('user__faculty__name').annotate(
                    count=Count('user', distinct=True)
                ).exclude(user__faculty__name__isnull=True)
                for fac_item in faculty_enrollments:
                    if fac_item['user__faculty__name']:
                        faculty_breakdown[fac_item['user__faculty__name']] = fac_item['count']
                
                # Reviews stats
                reviews = act.reviews.all()
                avg_rating = reviews.aggregate(avg=Avg('rating'))['avg'] or 0
                total_reviews = reviews.count()
                
                # Participation rate: users who actually participated vs enrolled
                users_who_participated = Participation.objects.filter(activity=act).exclude(user__id__in=admin_user_ids).values('user').distinct().count()
                participation_rate = 0
                if participants > 0:
                    participation_rate = round((users_who_participated / participants) * 100, 1) if users_who_participated > 0 else 0
                
                data[act.name] = {
                    'participacion': participants,  # Unique users enrolled
                    'actividades': enrollments.count(),  # Total enrollments
                    'participaciones_reales': participations,  # Actual participations/attendance
                    'tipo': act.type or 'N/A',
                    'categoria': act.category or 'N/A',
                    'ubicacion': act.location or 'N/A',
                    'capacidad_actual': participants,
                    'capacidad_maxima': act.max_capacity or 'Sin límite',
                    'tasa_participacion': f"{participation_rate}%",
                    'rating_promedio': round(avg_rating, 1) if avg_rating > 0 else 'N/A',
                    'total_resenas': total_reviews,
                    'distribucion_genero': gender_breakdown,
                    'distribucion_facultad': faculty_breakdown,
                    'requiere_inscripcion': 'Sí' if act.requires_registration else 'No',
                    'publicada': 'Sí' if act.is_published else 'No',
                }

        # 🔹 Filtro por facultad
        elif selected_filter == 'facultad':
            faculties = Faculty.objects.all().prefetch_related('users', 'users__enrollments', 
                                                               'users__enrollments__activity')
            for fac in faculties:
                users_in_faculty = fac.users.exclude(id__in=admin_user_ids)
                total_users = users_in_faculty.count()
                
                # Enrollments (exclude admin users)
                enrollments = Enrollment.objects.filter(user__faculty=fac).exclude(user__id__in=admin_user_ids)
                total_enrollments = enrollments.count()
                
                # Participations (exclude admin users)
                participations = Participation.objects.filter(user__faculty=fac).exclude(user__id__in=admin_user_ids).count()
                
                # Activities (distinct)
                activities_enrolled = enrollments.values('activity__name').distinct().count()
                
                # Gender breakdown
                gender_breakdown = {}
                for gender_code in ['M', 'F', 'O']:
                    gender_count = users_in_faculty.filter(gender=gender_code).count()
                    if gender_count > 0:
                        gender_breakdown[GENDER_NAMES.get(gender_code, gender_code)] = gender_count
                
                # Activity type breakdown
                activity_type_breakdown = {}
                type_enrollments = enrollments.values('activity__type').annotate(
                    count=Count('id')
                ).exclude(activity__type__isnull=True)
                for type_item in type_enrollments:
                    if type_item['activity__type']:
                        activity_type_breakdown[type_item['activity__type']] = type_item['count']
                
                # Top activities
                top_activities = enrollments.values('activity__name').annotate(
                    count=Count('user', distinct=True)
                ).order_by('-count')[:5]
                
                # Reviews given by users in this faculty (exclude admin users)
                reviews = ActivityReview.objects.filter(user__faculty=fac).exclude(user__id__in=admin_user_ids)
                avg_rating_given = reviews.aggregate(avg=Avg('rating'))['avg'] or 0
                total_reviews = reviews.count()
                
                data[fac.name] = {
                    'participacion': total_users,  # Total users in faculty
                    'actividades': total_enrollments,  # Total enrollments
                    'inscripciones': total_enrollments,
                    'participaciones_reales': participations,  # Actual participations
                    'actividades_unicas': activities_enrolled,  # Distinct activities
                    'distribucion_genero': gender_breakdown,
                    'distribucion_tipo_actividad': activity_type_breakdown,
                    'top_actividades': [item['activity__name'] for item in top_activities],
                    'rating_promedio_dado': round(avg_rating_given, 1) if avg_rating_given > 0 else 'N/A',
                    'total_resenas': total_reviews,
                    'tasa_participacion': round((participations / total_enrollments * 100), 1) if total_enrollments > 0 else 0,
                }

        # 🔹 Filtro por género
        elif selected_filter == 'genero':
            for gender_code in ['M', 'F', 'O']:
                users_with_gender = CustomUser.objects.filter(gender=gender_code).exclude(id__in=admin_user_ids)
                total_users = users_with_gender.count()
                
                if total_users == 0:
                    continue
                
                # Enrollments (exclude admin users)
                enrollments = Enrollment.objects.filter(user__gender=gender_code).exclude(user__id__in=admin_user_ids)
                total_enrollments = enrollments.count()
                
                # Participations (exclude admin users)
                participations = Participation.objects.filter(user__gender=gender_code).exclude(user__id__in=admin_user_ids).count()
                
                # Activities (distinct)
                activities_enrolled = enrollments.values('activity__name').distinct().count()
                
                # Faculty breakdown
                faculty_breakdown = {}
                faculty_users = users_with_gender.values('faculty__name').annotate(
                    count=Count('id')
                ).exclude(faculty__name__isnull=True)
                for fac_item in faculty_users:
                    if fac_item['faculty__name']:
                        faculty_breakdown[fac_item['faculty__name']] = fac_item['count']
                
                # Activity type breakdown
                activity_type_breakdown = {}
                type_enrollments = enrollments.values('activity__type').annotate(
                    count=Count('id')
                ).exclude(activity__type__isnull=True)
                for type_item in type_enrollments:
                    if type_item['activity__type']:
                        activity_type_breakdown[type_item['activity__type']] = type_item['count']
                
                # Top activities
                top_activities = enrollments.values('activity__name').annotate(
                    count=Count('user', distinct=True)
                ).order_by('-count')[:5]
                
                # Reviews given (exclude admin users)
                reviews = ActivityReview.objects.filter(user__gender=gender_code).exclude(user__id__in=admin_user_ids)
                avg_rating_given = reviews.aggregate(avg=Avg('rating'))['avg'] or 0
                total_reviews = reviews.count()
                
                # Average enrollments per user
                avg_enrollments_per_user = round(total_enrollments / total_users, 2) if total_users > 0 else 0
                
                gender_name = GENDER_NAMES.get(gender_code, gender_code)
                data[gender_name] = {
                    'participacion': total_users,  # Total users with this gender
                    'actividades': total_enrollments,  # Total enrollments
                    'inscripciones': total_enrollments,
                    'participaciones_reales': participations,  # Actual participations
                    'actividades_unicas': activities_enrolled,  # Distinct activities
                    'distribucion_facultad': faculty_breakdown,
                    'distribucion_tipo_actividad': activity_type_breakdown,
                    'top_actividades': [item['activity__name'] for item in top_activities],
                    'rating_promedio_dado': round(avg_rating_given, 1) if avg_rating_given > 0 else 'N/A',
                    'total_resenas': total_reviews,
                    'promedio_inscripciones_por_usuario': avg_enrollments_per_user,
                    'tasa_participacion': round((participations / total_enrollments * 100), 1) if total_enrollments > 0 else 0,
                }

        # 🔹 Filtro por semestre
        elif selected_filter == 'semestre':
            # Obtener todos los años únicos de enrollments y participations
            enrollment_years = Enrollment.objects.exclude(registered_at__isnull=True).values_list('registered_at__year', flat=True).distinct()
            participation_years = Participation.objects.exclude(attendance_date__isnull=True).values_list('attendance_date__year', flat=True).distinct()
            
            # Combinar años únicos
            all_years = set(enrollment_years) | set(participation_years)
            
            # Generar todos los semestres posibles para los años encontrados
            semesters = set()
            for year in all_years:
                semesters.add(f"{year}-01")  # Primer semestre
                semesters.add(f"{year}-02")  # Segundo semestre
            
            # Ordenar semestres
            sorted_semesters = sorted(semesters, reverse=True)
            
            for semester_str in sorted_semesters:
                start_date, end_date = get_semester_dates(semester_str)
                if not start_date or not end_date:
                    continue
                
                # Enrollments en este semestre (basado en registered_at, exclude admin users)
                enrollments = Enrollment.objects.filter(
                    registered_at__date__gte=start_date,
                    registered_at__date__lte=end_date
                ).exclude(user__id__in=admin_user_ids)
                total_enrollments = enrollments.count()
                unique_users_enrolled = enrollments.values('user').distinct().count()
                
                # Participations en este semestre (basado en attendance_date, exclude admin users)
                participations = Participation.objects.filter(
                    attendance_date__gte=start_date,
                    attendance_date__lte=end_date
                ).exclude(user__id__in=admin_user_ids)
                total_participations = participations.count()
                unique_users_participated = participations.values('user').distinct().count()
                
                # Solo incluir semestres que tienen datos
                if total_enrollments == 0 and total_participations == 0:
                    continue
                
                # Activities con enrollments en este semestre
                activities_with_enrollments = enrollments.values('activity__name').distinct().count()
                
                # Gender breakdown de usuarios inscritos
                gender_breakdown = {}
                for gender_code in ['M', 'F', 'O']:
                    gender_count = enrollments.filter(user__gender=gender_code).values('user').distinct().count()
                    if gender_count > 0:
                        gender_breakdown[GENDER_NAMES.get(gender_code, gender_code)] = gender_count
                
                # Faculty breakdown
                faculty_breakdown = {}
                faculty_enrollments = enrollments.values('user__faculty__name').annotate(
                    count=Count('user', distinct=True)
                ).exclude(user__faculty__name__isnull=True)
                for fac_item in faculty_enrollments:
                    if fac_item['user__faculty__name']:
                        faculty_breakdown[fac_item['user__faculty__name']] = fac_item['count']
                
                # Activity type breakdown
                activity_type_breakdown = {}
                type_enrollments = enrollments.values('activity__type').annotate(
                    count=Count('id')
                ).exclude(activity__type__isnull=True)
                for type_item in type_enrollments:
                    if type_item['activity__type']:
                        activity_type_breakdown[type_item['activity__type']] = type_item['count']
                
                # Top activities
                top_activities = enrollments.values('activity__name').annotate(
                    count=Count('user', distinct=True)
                ).order_by('-count')[:5]
                
                # Participation rate
                participation_rate = round((unique_users_participated / unique_users_enrolled * 100), 1) if unique_users_enrolled > 0 else 0
                
                # Semester name for display
                semester_display = f"{semester_str.split('-')[0]} - Semestre {semester_str.split('-')[1]}"
                if semester_str.endswith('-01'):
                    semester_display = f"{semester_str.split('-')[0]} - Primer Semestre (Ene-Jun)"
                elif semester_str.endswith('-02'):
                    semester_display = f"{semester_str.split('-')[0]} - Segundo Semestre (Jul-Dic)"
                
                data[semester_display] = {
                    'semestre': semester_str,
                    'fecha_inicio': start_date.strftime('%d/%m/%Y'),
                    'fecha_fin': end_date.strftime('%d/%m/%Y'),
                    'participacion': unique_users_enrolled,  # Unique users enrolled
                    'actividades': total_enrollments,  # Total enrollments
                    'participaciones_reales': total_participations,  # Actual participations
                    'usuarios_participaron': unique_users_participated,  # Unique users who participated
                    'actividades_unicas': activities_with_enrollments,  # Distinct activities
                    'distribucion_genero': gender_breakdown,
                    'distribucion_facultad': faculty_breakdown,
                    'distribucion_tipo_actividad': activity_type_breakdown,
                    'top_actividades': [item['activity__name'] for item in top_activities],
                    'tasa_participacion': participation_rate,
                }

        # 🔹 Filtro por año
        elif selected_filter == 'año':
            # Obtener todos los años únicos de enrollments y participations
            enrollment_years = Enrollment.objects.exclude(registered_at__isnull=True).values_list('registered_at__year', flat=True).distinct()
            participation_years = Participation.objects.exclude(attendance_date__isnull=True).values_list('attendance_date__year', flat=True).distinct()
            
            # Combinar años únicos
            years = set(enrollment_years) | set(participation_years)
            
            # Ordenar años
            sorted_years = sorted(years, reverse=True)
            
            for year in sorted_years:
                start_date, end_date = get_year_dates(str(year))
                if not start_date or not end_date:
                    continue
                
                # Enrollments en este año (exclude admin users)
                enrollments = Enrollment.objects.filter(
                    registered_at__date__gte=start_date,
                    registered_at__date__lte=end_date
                ).exclude(user__id__in=admin_user_ids)
                total_enrollments = enrollments.count()
                unique_users_enrolled = enrollments.values('user').distinct().count()
                
                # Participations en este año (exclude admin users)
                participations = Participation.objects.filter(
                    attendance_date__gte=start_date,
                    attendance_date__lte=end_date
                ).exclude(user__id__in=admin_user_ids)
                total_participations = participations.count()
                unique_users_participated = participations.values('user').distinct().count()
                
                # Solo incluir años que tienen datos
                if total_enrollments == 0 and total_participations == 0:
                    continue
                
                # Activities con enrollments en este año
                activities_with_enrollments = enrollments.values('activity__name').distinct().count()
                
                # Gender breakdown
                gender_breakdown = {}
                for gender_code in ['M', 'F', 'O']:
                    gender_count = enrollments.filter(user__gender=gender_code).values('user').distinct().count()
                    if gender_count > 0:
                        gender_breakdown[GENDER_NAMES.get(gender_code, gender_code)] = gender_count
                
                # Faculty breakdown
                faculty_breakdown = {}
                faculty_enrollments = enrollments.values('user__faculty__name').annotate(
                    count=Count('user', distinct=True)
                ).exclude(user__faculty__name__isnull=True)
                for fac_item in faculty_enrollments:
                    if fac_item['user__faculty__name']:
                        faculty_breakdown[fac_item['user__faculty__name']] = fac_item['count']
                
                # Activity type breakdown
                activity_type_breakdown = {}
                type_enrollments = enrollments.values('activity__type').annotate(
                    count=Count('id')
                ).exclude(activity__type__isnull=True)
                for type_item in type_enrollments:
                    if type_item['activity__type']:
                        activity_type_breakdown[type_item['activity__type']] = type_item['count']
                
                # Top activities
                top_activities = enrollments.values('activity__name').annotate(
                    count=Count('user', distinct=True)
                ).order_by('-count')[:5]
                
                # Breakdown por semestre dentro del año
                semester_breakdown = {}
                for semester_num in [1, 2]:
                    sem_start, sem_end = get_semester_dates(f"{year}-0{semester_num}")
                    if sem_start and sem_end:
                        sem_enrollments = enrollments.filter(
                            registered_at__date__gte=sem_start,
                            registered_at__date__lte=sem_end
                        ).count()
                        sem_participations = participations.filter(
                            attendance_date__gte=sem_start,
                            attendance_date__lte=sem_end
                        ).count()
                        sem_name = f"Semestre {semester_num}"
                        semester_breakdown[sem_name] = {
                            'inscripciones': sem_enrollments,
                            'participaciones': sem_participations,
                        }
                
                # Participation rate
                participation_rate = round((unique_users_participated / unique_users_enrolled * 100), 1) if unique_users_enrolled > 0 else 0
                
                data[str(year)] = {
                    'año': year,
                    'fecha_inicio': start_date.strftime('%d/%m/%Y'),
                    'fecha_fin': end_date.strftime('%d/%m/%Y'),
                    'participacion': unique_users_enrolled,  # Unique users enrolled
                    'actividades': total_enrollments,  # Total enrollments
                    'participaciones_reales': total_participations,  # Actual participations
                    'usuarios_participaron': unique_users_participated,  # Unique users who participated
                    'actividades_unicas': activities_with_enrollments,  # Distinct activities
                    'distribucion_genero': gender_breakdown,
                    'distribucion_facultad': faculty_breakdown,
                    'distribucion_tipo_actividad': activity_type_breakdown,
                    'distribucion_semestre': semester_breakdown,
                    'top_actividades': [item['activity__name'] for item in top_activities],
                    'tasa_participacion': participation_rate,
                }

        # Lista de filtros disponibles para el select en la plantilla
        filters = ['actividad', 'facultad', 'genero', 'semestre', 'año']

        context.update({
            'data': data,
            'filters': filters,
            'selected_filter': selected_filter,
        })

        return context


# =============================================================
# Vista para Reportes Formales de Participación
# =============================================================

class ParticipationFormalReportView(TemplateView):
    """
    Vista para generar reportes formales de participación con:
    - Filtros avanzados (frecuencia, tipo de actividad, programa/facultad, género)
    - Gráficos dinámicos
    - Cálculos: frecuencia promedio, actividad más popular, participación por grupo/horario
    - Exportación a CSV
    """
    template_name = 'reportsAndStats/participation_formal_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters from GET request
        # Note: getlist() returns empty list if no values selected, which is correct
        activity_types = [x for x in self.request.GET.getlist('activity_type') if x]  # Filter out empty strings
        faculties = [x for x in self.request.GET.getlist('faculty') if x]  # Filter out empty strings
        genders = [x for x in self.request.GET.getlist('gender') if x]  # Filter out empty strings
        frequency_min = self.request.GET.get('frequency_min', '').strip()  # Empty means no filter
        frequency_max = self.request.GET.get('frequency_max', '').strip()  # Empty means no filter
        date_start = self.request.GET.get('date_start', '').strip()
        date_end = self.request.GET.get('date_end', '').strip()
        
        # WORK LIKE FilteredReportsView: Use Enrollment as primary data source (since seed creates enrollments)
        # Also include Participation data if it exists
        
        # Exclude admin/CADI users from reports
        admin_user_ids = get_admin_user_ids()
        
        # Get base enrollments (primary data source - exists in seed)
        # Exclude admin users
        enrollments = Enrollment.objects.exclude(user__id__in=admin_user_ids).select_related(
            'user', 'activity', 'schedule', 'user__faculty'
        )
        
        # Get base participations (actual attendance - may not exist in seed)
        # Exclude admin users
        # NOTE: We filter for attendance_date__isnull=False to get only actual attendances
        # But if no participations exist, the report should still work with enrollments
        participations = Participation.objects.filter(
            attendance_date__isnull=False
        ).exclude(user__id__in=admin_user_ids).select_related('user', 'activity', 'schedule', 'user__faculty')
        
        # Store original unfiltered counts for debugging
        original_enrollment_count = enrollments.count()
        original_participation_count = participations.count()
        
        # Apply activity type filter
        if activity_types and len(activity_types) > 0:
            enrollments = enrollments.filter(activity__type__in=activity_types)
            participations = participations.filter(activity__type__in=activity_types)
        
        # Apply faculty filter
        if faculties and len(faculties) > 0:
            enrollments = enrollments.filter(user__faculty__name__in=faculties)
            participations = participations.filter(user__faculty__name__in=faculties)
        
        # Apply gender filter
        if genders and len(genders) > 0:
            enrollments = enrollments.filter(user__gender__in=genders)
            participations = participations.filter(user__gender__in=genders)
        
        # Apply date filters (only if date strings are not empty)
        if date_start and date_start.strip():
            try:
                start_date = datetime.strptime(date_start.strip(), '%Y-%m-%d').date()
                enrollments = enrollments.filter(registered_at__date__gte=start_date)
                participations = participations.filter(attendance_date__gte=start_date)
            except (ValueError, TypeError) as e:
                # If date parsing fails, ignore this filter
                pass
        
        if date_end and date_end.strip():
            try:
                end_date = datetime.strptime(date_end.strip(), '%Y-%m-%d').date()
                enrollments = enrollments.filter(registered_at__date__lte=end_date)
                participations = participations.filter(attendance_date__lte=end_date)
            except (ValueError, TypeError) as e:
                # If date parsing fails, ignore this filter
                pass
        
        # For frequency calculation: count enrollments per user (engagement frequency)
        # Calculate frequency based on CURRENT filtered enrollments
        # This ensures frequency is calculated relative to the filtered dataset
        user_enrollment_counts = list(enrollments.values('user').annotate(
            enrollment_count=Count('id')
        ).order_by('-enrollment_count'))
        
        # Apply frequency filter based on ENROLLMENT count
        # Frequency filter: show only users whose enrollment count (within current filters) matches the range
        # Only apply if at least one frequency value is provided and not empty
        freq_min_provided = frequency_min and frequency_min.strip()
        freq_max_provided = frequency_max and frequency_max.strip()
        
        if freq_min_provided or freq_max_provided:
            try:
                # Parse frequency values - use 0 and very large number as defaults if not provided
                freq_min = int(frequency_min.strip()) if freq_min_provided else 0
                freq_max = int(frequency_max.strip()) if freq_max_provided else 999999
                
                # Validate range
                if freq_min < 0:
                    freq_min = 0
                if freq_max < freq_min:
                    freq_max = freq_min
                
                if user_enrollment_counts:
                    user_ids_with_frequency = [
                        item['user'] for item in user_enrollment_counts
                        if item['user'] is not None and freq_min <= item['enrollment_count'] <= freq_max
                    ]
                    if user_ids_with_frequency:
                        # Apply frequency filter to both enrollments and participations
                        enrollments = enrollments.filter(user__id__in=user_ids_with_frequency)
                        participations = participations.filter(user__id__in=user_ids_with_frequency)
                    else:
                        # No users match the frequency range - return empty querysets
                        enrollments = Enrollment.objects.none()
                        participations = Participation.objects.none()
            except (ValueError, TypeError) as e:
                # If frequency filter is invalid, ignore it and continue with all data
                pass
        
        # ========== CALCULATIONS ==========
        # Use enrollments as primary data source (like FilteredReportsView) since seed creates enrollments
        # Also include participations when they exist
        
        # Debug: Check data availability (excluding admin users, but before applying filters)
        total_enrollments_in_db = Enrollment.objects.exclude(user__id__in=admin_user_ids).count()
        total_participations_in_db = Participation.objects.filter(
            attendance_date__isnull=False
        ).exclude(user__id__in=admin_user_ids).count()
        
        # Use enrollments for user count and engagement metrics (like FilteredReportsView)
        # This ensures it works with seed data which has enrollments
        total_users = enrollments.values('user').distinct().count()
        total_enrollments = enrollments.count()
        
        # Use participations for actual attendance metrics
        total_participations = participations.count()
        
        # Average frequency: average enrollments per user (engagement frequency)
        avg_frequency = round(total_enrollments / total_users, 2) if total_users > 0 else 0
        
        # Average participation frequency: average actual participations per user
        avg_participation_frequency = round(total_participations / total_users, 2) if total_users > 0 else 0
        
        # 2. Most popular activity (by enrollment count - like FilteredReportsView)
        # Use enrollments since that's what seed data has
        most_popular_activity = None
        if total_enrollments > 0:
            most_popular_activity_data = enrollments.values('activity__name', 'activity__type').annotate(
                enrollment_count=Count('id')
            ).order_by('-enrollment_count').first()
            if most_popular_activity_data:
                # Also get participation count for this activity
                activity_name = most_popular_activity_data['activity__name']
                participation_count = participations.filter(activity__name=activity_name).count()
                most_popular_activity = {
                    'activity__name': activity_name,
                    'activity__type': most_popular_activity_data['activity__type'],
                    'enrollment_count': most_popular_activity_data['enrollment_count'],
                    'participation_count': participation_count,
                }
        
        # 3. Participation by schedule/group (grouped by activity and schedule)
        # Use enrollments with schedule info (like FilteredReportsView)
        enrollment_by_schedule = list(enrollments.values(
            'activity__name',
            'activity__type',
            'schedule__day',
            'schedule__start_time',
            'schedule__end_time'
        ).annotate(
            enrollment_count=Count('id'),
            unique_users=Count('user', distinct=True)
        ).order_by('-enrollment_count'))
        
        # Also get participation counts by schedule
        participation_by_schedule = []
        if total_participations > 0:
            participation_by_schedule = list(participations.values(
                'activity__name',
                'activity__type',
                'schedule__day',
                'schedule__start_time',
                'schedule__end_time'
            ).annotate(
                participation_count=Count('id'),
                unique_users=Count('user', distinct=True)
            ).order_by('-participation_count'))
        
        # Group by activity for schedule breakdown
        # Combine enrollment and participation data
        schedule_breakdown = {}
        
        # First, add enrollment data (primary source)
        for item in enrollment_by_schedule:
            activity_name = item['activity__name'] or 'Sin actividad'
            schedule_key = f"{item['schedule__day'] or 'N/A'} {item['schedule__start_time'] or ''}-{item['schedule__end_time'] or ''}"
            
            if activity_name not in schedule_breakdown:
                schedule_breakdown[activity_name] = {
                    'activity_type': item['activity__type'] or 'N/A',
                    'schedules': [],
                    'total_enrollments': 0,
                    'total_participations': 0,
                    'total_users': 0,
                }
            
            # Get participation count for this schedule
            participation_count = 0
            if participation_by_schedule:
                for p_item in participation_by_schedule:
                    if (p_item['activity__name'] == activity_name and 
                        p_item['schedule__day'] == item['schedule__day'] and
                        p_item['schedule__start_time'] == item['schedule__start_time']):
                        participation_count = p_item['participation_count']
                        break
            
            schedule_breakdown[activity_name]['schedules'].append({
                'schedule': schedule_key,
                'enrollment_count': item['enrollment_count'],
                'participation_count': participation_count,
                'unique_users': item['unique_users'],
            })
            schedule_breakdown[activity_name]['total_enrollments'] += item['enrollment_count']
            schedule_breakdown[activity_name]['total_participations'] += participation_count
            schedule_breakdown[activity_name]['total_users'] = max(
                schedule_breakdown[activity_name]['total_users'],
                item['unique_users']
            )
        
        # 4. Engagement by activity type (use enrollments like FilteredReportsView)
        enrollment_by_type = list(enrollments.values('activity__type').annotate(
            enrollment_count=Count('id'),
            unique_users=Count('user', distinct=True),
            unique_activities=Count('activity', distinct=True)
        ).order_by('-enrollment_count'))
        
        # Also get participation counts by type
        participation_by_type_list = list(participations.values('activity__type').annotate(
            participation_count=Count('id'),
            unique_users=Count('user', distinct=True),
            unique_activities=Count('activity', distinct=True)
        ).order_by('-participation_count'))
        
        # Combine for display (prefer enrollment data, add participation data)
        activity_type_data = {}
        # Start with enrollment data (primary source)
        for item in enrollment_by_type:
            if item['activity__type']:  # Only add if type is not None/empty
                activity_type_data[item['activity__type']] = {
                    'enrollment_count': item['enrollment_count'],
                    'participation_count': 0,
                    'unique_users': item['unique_users'],
                    'unique_activities': item['unique_activities'],
                }
        # Add participation counts to existing types
        for item in participation_by_type_list:
            if item['activity__type'] and item['activity__type'] in activity_type_data:
                activity_type_data[item['activity__type']]['participation_count'] = item['participation_count']
        # Also add participation-only types (in case there are participations without enrollments)
        for item in participation_by_type_list:
            if item['activity__type'] and item['activity__type'] not in activity_type_data:
                activity_type_data[item['activity__type']] = {
                    'enrollment_count': 0,
                    'participation_count': item['participation_count'],
                    'unique_users': item['unique_users'],
                    'unique_activities': item['unique_activities'],
                }
        
        # 5. Engagement by faculty (use enrollments)
        enrollment_by_faculty = list(enrollments.values('user__faculty__name').annotate(
            enrollment_count=Count('id'),
            unique_users=Count('user', distinct=True)
        ).exclude(user__faculty__name__isnull=True).order_by('-enrollment_count'))
        
        participation_by_faculty_list = list(participations.values('user__faculty__name').annotate(
            participation_count=Count('id'),
            unique_users=Count('user', distinct=True)
        ).exclude(user__faculty__name__isnull=True).order_by('-participation_count'))
        
        # 6. Engagement by gender (use enrollments)
        enrollment_by_gender = list(enrollments.values('user__gender').annotate(
            enrollment_count=Count('id'),
            unique_users=Count('user', distinct=True)
        ).exclude(user__gender__isnull=True).order_by('-enrollment_count'))
        
        participation_by_gender_list = list(participations.values('user__gender').annotate(
            participation_count=Count('id'),
            unique_users=Count('user', distinct=True)
        ).exclude(user__gender__isnull=True).order_by('-participation_count'))
        
        GENDER_NAMES = {
            'M': 'Masculino',
            'F': 'Femenino',
            'O': 'Otro',
        }
        
        # 7. Frequency distribution (based on enrollment count - like FilteredReportsView)
        frequency_distribution = {}
        if total_enrollments > 0:
            user_enrollment_frequencies = enrollments.values('user').annotate(
                enrollment_count=Count('id')
            )
            for item in user_enrollment_frequencies:
                count = item['enrollment_count']
                frequency_range = self._get_frequency_range(count)
                if frequency_range not in frequency_distribution:
                    frequency_distribution[frequency_range] = 0
                frequency_distribution[frequency_range] += 1
        
        # 8. Top activities by enrollment (like FilteredReportsView)
        top_activities = []
        if total_enrollments > 0:
            top_activities_enrollments = list(enrollments.values(
                'activity__name',
                'activity__type'
            ).annotate(
                enrollment_count=Count('id'),
                unique_users=Count('user', distinct=True)
            ).order_by('-enrollment_count')[:10])
            
            # Add participation counts to top activities
            for item in top_activities_enrollments:
                activity_name = item['activity__name']
                participation_count = participations.filter(activity__name=activity_name).count()
                top_activities.append({
                    'activity__name': activity_name,
                    'activity__type': item['activity__type'],
                    'enrollment_count': item['enrollment_count'],
                    'participation_count': participation_count,
                    'unique_users': item['unique_users'],
                })
        
        # Prepare data for charts (use enrollment data primarily, like FilteredReportsView)
        chart_data = {
            'by_type': [
                {
                    'type': activity_type or 'N/A',
                    'count': data['enrollment_count'],  # Use enrollment count
                    'participation_count': data['participation_count'],  # Also include participation
                    'users': data['unique_users'],
                    'activities': data['unique_activities']
                }
                for activity_type, data in activity_type_data.items()
            ],
            'by_faculty': [
                {
                    'faculty': item['user__faculty__name'] or 'N/A',
                    'count': item['enrollment_count'],  # Use enrollment count
                    'users': item['unique_users']
                }
                for item in enrollment_by_faculty
            ],
            'by_gender': [
                {
                    'gender': GENDER_NAMES.get(item['user__gender'], item['user__gender'] or 'N/A'),
                    'count': item['enrollment_count'],  # Use enrollment count
                    'users': item['unique_users']
                }
                for item in enrollment_by_gender
            ],
            'frequency_distribution': frequency_distribution,
            'top_activities': [
                {
                    'name': item['activity__name'] or 'N/A',
                    'type': item['activity__type'] or 'N/A',
                    'count': item['enrollment_count'],  # Use enrollment count
                    'participation_count': item.get('participation_count', 0),
                    'users': item['unique_users']
                }
                for item in top_activities
            ],
        }
        
        # Get available options for filters
        available_activity_types = Activity.objects.values_list('type', flat=True).distinct().exclude(type__isnull=True)
        # Exclude CADI from faculty filter options (CADI is admin, not a participant faculty)
        available_faculties = Faculty.objects.values_list('name', flat=True).distinct().exclude(name='CADI')
        
        context.update({
            'avg_frequency': avg_frequency,
            'avg_participation_frequency': avg_participation_frequency,
            'most_popular_activity': most_popular_activity,
            'schedule_breakdown': schedule_breakdown,
            'participation_by_type': list(activity_type_data.items()),  # Convert to list for template
            'participation_by_faculty': enrollment_by_faculty,
            'participation_by_gender': enrollment_by_gender,
            'frequency_distribution': frequency_distribution,
            'top_activities': top_activities,
            'total_participations': total_participations,
            'total_enrollments': total_enrollments,
            'total_users': total_users,
            'total_activities': enrollments.values('activity').distinct().count(),
            'chart_data_json': json.dumps(chart_data),
            'available_activity_types': available_activity_types,
            'available_faculties': available_faculties,
            'selected_activity_types': activity_types,
            'selected_faculties': faculties,
            'selected_genders': genders,
            'frequency_min': frequency_min if frequency_min else '',
            'frequency_max': frequency_max if frequency_max else '',
            'date_start': date_start,
            'date_end': date_end,
            'gender_names': GENDER_NAMES,
            'debug_total_in_db': total_enrollments_in_db,  # Show enrollments in DB
            'debug_participations_in_db': total_participations_in_db,  # Show participations in DB
            'debug_original_enrollment_count': original_enrollment_count,  # Before filters
            'debug_original_participation_count': original_participation_count,  # Before filters
        })
        
        return context
    
    def _get_frequency_range(self, count):
        """Categorize participation frequency into ranges"""
        if count == 0:
            return '0'
        elif count <= 5:
            return '1-5'
        elif count <= 10:
            return '6-10'
        elif count <= 20:
            return '11-20'
        elif count <= 50:
            return '21-50'
        else:
            return '50+'


class ParticipationReportExportView(View):
    """
    View to export participation report data to CSV
    """
    def get(self, request):
        # Get same filters as the report view (use same logic)
        activity_types = request.GET.getlist('activity_type')
        faculties = request.GET.getlist('faculty')
        genders = request.GET.getlist('gender')
        frequency_min = request.GET.get('frequency_min', '').strip()  # Empty means no filter
        frequency_max = request.GET.get('frequency_max', '').strip()  # Empty means no filter
        date_start = request.GET.get('date_start', '').strip()
        date_end = request.GET.get('date_end', '').strip()
        
        # Exclude admin/CADI users from export
        admin_user_ids = get_admin_user_ids()
        
        # Use SAME logic as ParticipationFormalReportView
        # Filter out empty strings from lists
        activity_types = [x for x in activity_types if x]
        faculties = [x for x in faculties if x]
        genders = [x for x in genders if x]
        frequency_min = frequency_min.strip() if frequency_min else ''
        frequency_max = frequency_max.strip() if frequency_max else ''
        date_start = date_start.strip() if date_start else ''
        date_end = date_end.strip() if date_end else ''
        
        # Get enrollments (primary data source - like report view)
        enrollments = Enrollment.objects.exclude(user__id__in=admin_user_ids).select_related(
            'user', 'activity', 'schedule', 'user__faculty'
        )
        
        # Get participations (actual attendance)
        participations = Participation.objects.filter(
            attendance_date__isnull=False
        ).exclude(user__id__in=admin_user_ids).select_related('user', 'activity', 'schedule', 'user__faculty')
        
        # Apply filters to both (same as report view)
        if activity_types and len(activity_types) > 0:
            enrollments = enrollments.filter(activity__type__in=activity_types)
            participations = participations.filter(activity__type__in=activity_types)
        
        if faculties and len(faculties) > 0:
            enrollments = enrollments.filter(user__faculty__name__in=faculties)
            participations = participations.filter(user__faculty__name__in=faculties)
        
        if genders and len(genders) > 0:
            enrollments = enrollments.filter(user__gender__in=genders)
            participations = participations.filter(user__gender__in=genders)
        
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
                enrollments = enrollments.filter(registered_at__date__gte=start_date)
                participations = participations.filter(attendance_date__gte=start_date)
            except (ValueError, TypeError):
                pass
        
        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
                enrollments = enrollments.filter(registered_at__date__lte=end_date)
                participations = participations.filter(attendance_date__lte=end_date)
            except (ValueError, TypeError):
                pass
        
        # Apply frequency filter (based on enrollment count, like report view)
        freq_min_provided = frequency_min and frequency_min.strip()
        freq_max_provided = frequency_max and frequency_max.strip()
        
        if freq_min_provided or freq_max_provided:
            try:
                # Calculate frequency based on enrollments
                user_enrollment_counts = list(enrollments.values('user').annotate(
                    enrollment_count=Count('id')
                ))
                
                freq_min = int(frequency_min.strip()) if freq_min_provided else 0
                freq_max = int(frequency_max.strip()) if freq_max_provided else 999999
                
                if freq_min < 0:
                    freq_min = 0
                if freq_max < freq_min:
                    freq_max = freq_min
                
                if user_enrollment_counts:
                    user_ids_with_frequency = [
                        item['user'] for item in user_enrollment_counts
                        if item['user'] is not None and freq_min <= item['enrollment_count'] <= freq_max
                    ]
                    if user_ids_with_frequency:
                        enrollments = enrollments.filter(user__id__in=user_ids_with_frequency)
                        participations = participations.filter(user__id__in=user_ids_with_frequency)
                    else:
                        # No users match - return empty querysets
                        enrollments = Enrollment.objects.none()
                        participations = Participation.objects.none()
            except (ValueError, TypeError):
                pass
        
        # For CSV export, we want to export BOTH enrollments and participations
        # But prioritize participations (actual attendance) if they exist
        # If no participations, export enrollment data
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="reporte_participacion.csv"'
        
        # Write BOM for Excel compatibility with UTF-8
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'ID Estudiante',
            'Nombre',
            'Facultad',
            'Género',
            'Actividad',
            'Tipo Actividad',
            'Horario',
            'Fecha Registro/Asistencia',
            'Hora',
            'Tipo',
        ])
        
        # GENDER_NAMES mapping
        GENDER_NAMES = {
            'M': 'Masculino',
            'F': 'Femenino',
            'O': 'Otro',
        }
        
        rows_written = 0
        
        # Export participations (actual attendance) if they exist
        for participation in participations.order_by('user__identification', 'attendance_date'):
            schedule_str = 'N/A'
            if participation.schedule:
                schedule_str = f"{participation.schedule.day} {participation.schedule.start_time}-{participation.schedule.end_time}"
            
            # Get gender display
            gender_display = GENDER_NAMES.get(participation.user.gender, participation.user.gender or 'N/A')
            
            writer.writerow([
                participation.user.identification or 'N/A',
                participation.user.get_full_name() or participation.user.username,
                participation.user.faculty.name if participation.user.faculty else 'N/A',
                gender_display,
                participation.activity.name,
                participation.activity.type or 'N/A',
                schedule_str,
                participation.attendance_date.strftime('%Y-%m-%d') if participation.attendance_date else 'N/A',
                participation.attendance_time.strftime('%H:%M:%S') if participation.attendance_time else 'N/A',
                'Participación',
            ])
            rows_written += 1
        
        # If no participations or very few, also export enrollments (inscriptions)
        # This ensures CSV has data when seed only has enrollments
        if rows_written == 0:
            # Export enrollments when there are no participations
            for enrollment in enrollments.order_by('user__identification', 'registered_at'):
                schedule_str = 'N/A'
                if enrollment.schedule:
                    schedule_str = f"{enrollment.schedule.day} {enrollment.schedule.start_time}-{enrollment.schedule.end_time}"
                
                # Get gender display
                gender_display = GENDER_NAMES.get(enrollment.user.gender, enrollment.user.gender or 'N/A')
                
                writer.writerow([
                    enrollment.user.identification or 'N/A',
                    enrollment.user.get_full_name() or enrollment.user.username,
                    enrollment.user.faculty.name if enrollment.user.faculty else 'N/A',
                    gender_display,
                    enrollment.activity.name,
                    enrollment.activity.type or 'N/A',
                    schedule_str,
                    enrollment.registered_at.strftime('%Y-%m-%d') if enrollment.registered_at else 'N/A',
                    enrollment.registered_at.strftime('%H:%M:%S') if enrollment.registered_at else 'N/A',
                    'Inscripción',
                ])
                rows_written += 1
        
        return response