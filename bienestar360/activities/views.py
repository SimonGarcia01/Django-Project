import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.views.generic import (
    CreateView, UpdateView, DeleteView, ListView, DetailView,
    TemplateView, View
)
from django import forms
from collections import defaultdict
from django.db.models import Avg
from django.urls import reverse, reverse_lazy
from django.db.models.functions import TruncDate, TruncMonth
from django.db.models import Count
from django.utils import timezone
from datetime import timedelta, datetime
import csv
import re
import calendar
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib.auth.models import Group
from django.db.models import Q
from login.models import CustomUser
from .models import Activity, Enrollment, ActivityReview, Schedule
from django.db import IntegrityError
from social_projects.models import SocialProject, SocialEvent, SocialEventEnrollment
from datetime import datetime
from .models import Activity, Enrollment, ActivityReview, Schedule, Participation
from django.db import IntegrityError
import calendar
from social_projects.models import SocialProject
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Email stuff
from django.core.signing import dumps, loads, BadSignature, SignatureExpired
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings

from .models import Activity, Enrollment, ActivityReview, Schedule, Participation, Evento
from .forms import ActivityForm, ScheduleForm
from tournaments.models import Tournament, TournamentGame


# ============================================================== #
# Permission helpers
# ==============================================================

def is_staff_user(user):
    """Check if user is staff (admin group or superuser)"""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    # Check if user belongs to admin group
    return user.groups.filter(name='admin').exists()


class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin to require staff permissions"""
    def test_func(self):
        return is_staff_user(self.request.user)
    
    def handle_no_permission(self):
        messages.error(self.request, "No tienes permiso para acceder a esta página.")
        return redirect('login:login')

# Shared constants
weekday_map = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}


# ============================================================== #
# CRUD de Actividades (Administrador)
# ============================================================== #

class CreateActivityView(StaffRequiredMixin, CreateView):
    model = Activity
    form_class = ActivityForm
    template_name = "activities/create_activity.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ScheduleFormSet = forms.formset_factory(ScheduleForm, extra=1)
        if "formset" not in kwargs:
            if self.request.method == "POST":
                context["formset"] = ScheduleFormSet(self.request.POST, prefix="schedules")
            else:
                context["formset"] = ScheduleFormSet(prefix="schedules")
        else:
            context["formset"] = kwargs["formset"]
        return context

    def form_valid(self, form):
        ScheduleFormSet = forms.formset_factory(ScheduleForm, extra=1)
        formset = ScheduleFormSet(self.request.POST, prefix="schedules")

        if formset.is_valid():
            activity = form.save()
            is_evento = activity.type == "Eventos"
            for schedule_form in formset:
                day = schedule_form.cleaned_data.get("day")
                start_time = schedule_form.cleaned_data.get("start_time")
                end_time = schedule_form.cleaned_data.get("end_time")
                
                if is_evento and not day and activity.description:
                    import re
                    match = re.search(r'FECHA:(\d{4}-\d{2}-\d{2})', activity.description)
                    if match:
                        try:
                            fecha_str = match.group(1)
                            evento_fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
                            day = weekday_map.get(evento_fecha.weekday(), "Lunes")
                        except ValueError:
                            day = "Lunes"
                    else:
                        day = "Lunes"
                
                if is_evento:
                    if start_time and end_time:
                        if not day:
                            day = "Lunes"
                        Schedule.objects.create(
                            activity=activity,
                            day=day,
                            start_time=start_time,
                            end_time=end_time
                        )
                else:
                    if day and start_time and end_time:
                        Schedule.objects.create(
                            activity=activity,
                            day=day,
                            start_time=start_time,
                            end_time=end_time
                        )
            messages.success(self.request, "La actividad se creó correctamente.")
            return redirect("public_activities")
        else:
            # If formset is invalid, return form_invalid to show errors
            return self.render_to_response(self.get_context_data(form=form, formset=formset))


class UpdateActivityView(StaffRequiredMixin, UpdateView):
    model = Activity
    form_class = ActivityForm
    template_name = "activities/update_activity.html"
    pk_url_kwarg = "pk"
    success_url = reverse_lazy("public_activities")

    def form_valid(self, form):
        response = super().form_valid(form)
        # Save the form
        self.object = form.save()
        messages.success(self.request, "La actividad se actualizó correctamente.")
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["activity"] = self.get_object()
        return context


class DeleteActivityView(StaffRequiredMixin, DeleteView):
    model = Activity
    pk_url_kwarg = "pk"
    success_url = None

    def get_success_url(self):
        return reverse("public_activities")

    def delete(self, request, *args, **kwargs):
        messages.info(request, "La actividad ha sido eliminada.")
        return super().delete(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        # Override GET to delete immediately (since we have JavaScript confirmation)
        # This allows the link to work directly
        self.object = self.get_object()
        self.object.delete()
        messages.info(request, "La actividad ha sido eliminada.")
        return redirect(self.get_success_url())


class PublicActivitiesListView(StaffRequiredMixin, ListView):
    model = Activity
    template_name = "activities/public_activities.html"
    context_object_name = "activities"

    def get_queryset(self):
        queryset = Activity.objects.filter(is_published=True)
        
        activity_type = self.request.GET.get("type")
        location = self.request.GET.get("location")
        time = self.request.GET.get("time")

        if activity_type:
            queryset = queryset.filter(type=activity_type)
        if location:
            queryset = queryset.filter(location__icontains=location)
        if time:
            queryset = queryset.filter(
                schedules__start_time__lte=time,
                schedules__end_time__gte=time
            ).distinct()
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["selected_type"] = self.request.GET.get("type")
        context["selected_location"] = self.request.GET.get("location")
        context["selected_time"] = self.request.GET.get("time")
        return context


# ============================================================== #
# Vista principal para estudiantes
# ============================================================== #

class ActivityView(ListView):
    model = Activity
    template_name = "activities/activityHomePage.html"
    context_object_name = "activities"

    def get_queryset(self):
        queryset = Activity.objects.filter(is_published=True)
        selected_type = self.request.GET.get("type")
        selected_location = self.request.GET.get("location")
        selected_time = self.request.GET.get("time")

        if selected_type:
            queryset = queryset.filter(type=selected_type)
        if selected_location:
            queryset = queryset.filter(location__icontains=selected_location)
        if selected_time:
            queryset = queryset.filter(
                schedules__start_time__lte=selected_time,
                schedules__end_time__gte=selected_time
            ).distinct()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["selected_type"] = self.request.GET.get("type")
        context["selected_location"] = self.request.GET.get("location")
        context["selected_time"] = self.request.GET.get("time")
        return context


# ============================================================== #
# Inscripción de estudiantes
# ============================================================== #

class ActivityListView(ListView):
    model = Activity
    template_name = "activities/activity_list.html"

    def get_queryset(self):
        return Activity.objects.filter(is_published=True)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        activities = self.get_queryset()
        context["no_registration"] = activities.filter(requires_registration=False)
        context["with_registration"] = activities.filter(requires_registration=True)
        return context


class ActivityDetailView(DetailView):
    model = Activity
    template_name = "activities/activity_detail.html"
    pk_url_kwarg = "activity_id"
    context_object_name = "activity"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        activity = self.get_object()
        context["schedules"] = activity.schedules.all().order_by("day", "start_time")
        context["current_registrations"] = activity.current_registrations_count
        context["max_capacity"] = activity.max_capacity
        context["is_full"] = activity.is_full_status
        return context


class EnrollInActivityView(LoginRequiredMixin, View):
    def get(self, request, activity_id):
        # Redirect GET requests to activity view
        return redirect("activityView")

    def post(self, request, activity_id):
        activity = get_object_or_404(Activity, pk=activity_id)
        if not activity.requires_registration:
            messages.warning(request, "Esta actividad no requiere inscripción.")
            return redirect("activityView")

        if activity.is_full_status:
            messages.error(request, "El cupo de esta actividad está lleno.")
            return redirect("activityView")

        enrollment, created = Enrollment.objects.get_or_create(
            user=request.user,
            activity=activity
        )

        if not created:
            messages.info(request, "Ya estabas inscrito en esta actividad.")
        else:
            messages.success(request, "¡Te inscribiste con éxito!")

            # Send an email if the activity requires registration
            try:
                # token with the minimum info to confirm
                token = dumps({"enrollment_id": enrollment.id, "user_id": request.user.id})
                confirm_path = reverse("confirm_enrollment", args=[token])
                confirm_url = request.build_absolute_uri(confirm_path)

                subject = f"Confirma tu asistencia a {activity.name}"
                message = render_to_string("activities/confirm_enrollment_email.txt", {
                    "user": request.user,
                    "activity": activity,
                    "confirm_url": confirm_url,
                })

                #There is a Default but it may not be set
                from_email = getattr(settings, "DEFAULT_FROM_EMAIL", None) or "no-reply@localhost"

                # send_mail will use the project's EMAIL_BACKEND
                send_mail(subject, message, from_email, [request.user.email], fail_silently=False)
            except Exception:
                # If the email fails, make a log of the error but don't block the enrollment
                pass

        return redirect("activityView")


class ConfirmEnrollmentView(View):
    """Vista para confirmar la asistencia desde el enlace enviado por correo.
    El token es un valor firmado que contiene 'enrollment_id' y 'user_id'.
    """
    def get(self, request, token):
        #Check any problems with the token
        try:
            data = loads(token, max_age=60 * 60 * 24 * 7)  #make a valid token for 7 days
        except SignatureExpired:
            messages.error(request, "El enlace de confirmación ha expirado.")
            return redirect("activityView")
        except BadSignature:
            messages.error(request, "Enlace de confirmación inválido.")
            return redirect("activityView")

        enrollment_id = data.get("enrollment_id")
        user_id = data.get("user_id")

        # Now actually check if the enrollment exists
        try:
            enrollment = Enrollment.objects.select_related("activity", "user").get(pk=enrollment_id)
        except Enrollment.DoesNotExist:
            messages.error(request, "Inscripción no encontrada.")
            return redirect("activityView")

        # Verify the if it coincides with the user and if the user is authenticated
        if request.user.is_authenticated and request.user.id != user_id:
            messages.error(request, "No tienes permiso para confirmar esta inscripción.")
            return redirect("activityView")

        # Now here actually confirm the enrollment if it passes all checks
        enrollment.confirm()
        messages.success(request, f"Has confirmado tu asistencia a '{enrollment.activity.name}'.")
        return redirect("my_calendar")


# ============================================================== #
# Reseñas de Actividades
# ============================================================== #

class ReviewActivityView(LoginRequiredMixin, View):
    """View for students to submit reviews/feedback on activities"""
    
    def get(self, request, pk):
        activity = get_object_or_404(Activity, pk=pk)
        return render(request, "activities/activity_review.html", {"activity": activity})

    def post(self, request, pk):
        activity = get_object_or_404(Activity, pk=pk)
        rating = request.POST.get("rating")
        comment = request.POST.get("comment", "")
        
        # Server-side validation
        errors = []
        
        # Validate rating is provided
        if not rating:
            errors.append("La valoración es requerida.")
        else:
            try:
                rating_int = int(rating)
                # Validate rating is between 1 and 5
                if rating_int < 1 or rating_int > 5:
                    errors.append("La valoración debe estar entre 1 y 5.")
                    rating = None
            except ValueError:
                errors.append("La valoración debe ser un número válido.")
                rating = None
        
        # If there are errors, show them and re-render the form
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, "activities/activity_review.html", {
                "activity": activity,
                "errors": errors
            })
        
        # Create the review
        if rating:
            ActivityReview.objects.create(
                activity=activity,
                user=request.user,
                rating=int(rating),
                comment=comment if comment else None
            )
            messages.success(request, "¡Gracias! Tu reseña ha sido enviada correctamente.")
            return redirect("activity_reviews", pk=pk)
        
        # Fallback: should not reach here if validation worked
        return render(request, "activities/activity_review.html", {"activity": activity})


class ActivityReviewsView(DetailView):
    model = Activity
    template_name = "activities/activity_reviews.html"
    pk_url_kwarg = "pk"
    context_object_name = "activity"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        activity = self.get_object()
        reviews = ActivityReview.objects.filter(activity=activity).select_related("user").order_by("-created_at")
        average_rating = reviews.aggregate(avg_rating=Avg("rating"))["avg_rating"] or 0
        context["reviews"] = reviews
        context["average_rating"] = round(average_rating, 1)
        return context


# ============================================================== #
# CADI: revisión de reseñas
# ============================================================== #

class CADIActivityReviewView(StaffRequiredMixin, View):
    """
    Vista del CADI: muestra todas las actividades con sus reseñas,
    o una sola actividad si se pasa el parámetro pk.
    Requiere que el usuario sea staff (admin).
    """
    def get(self, request, pk=None):
        if pk:
            activity = get_object_or_404(Activity, pk=pk)
            reviews = ActivityReview.objects.filter(activity=activity).select_related("user").order_by("-created_at")
            average_rating = reviews.aggregate(avg_rating=Avg("rating"))["avg_rating"] or 0
            return render(request, "activities/CADI_activity_review.html", {
                "single_activity": True,
                "activity": activity,
                "reviews": reviews,
                "average_rating": round(average_rating, 1),
            })
        else:
            # Mostrar todas las actividades con sus reseñas correspondientes
            activities = Activity.objects.filter(is_published=True).prefetch_related("reviews__user")
            data = []
            for activity in activities:
                reviews = activity.reviews.all().order_by("-created_at")
                avg = reviews.aggregate(avg_rating=Avg("rating"))["avg_rating"] or 0
                data.append({
                    "activity": activity,
                    "reviews": reviews,
                    "average_rating": round(avg, 1),
                })
            return render(request, "activities/CADI_activity_review.html", {
                "single_activity": False,
                "data": data,
            })


class MarkReviewReadView(StaffRequiredMixin, View):
    """
    Marca una ActivityReview como leída (is_read=True).
    - Requiere que el usuario sea staff.
    - Idealmente se invoca por POST; si se invoca por GET redirige igualmente (compatibilidad).
    """
    def get(self, request, review_id):
        return self._mark_read(request, review_id)

    def post(self, request, review_id):
        return self._mark_read(request, review_id)

    def _mark_read(self, request, review_id):
        review = get_object_or_404(ActivityReview, pk=review_id)
        review.is_read = True
        review.save()
        messages.success(request, "La reseña ha sido marcada como leída.")
        return redirect(reverse('cadi_activity_review_detail', args=[review.activity.pk]))


# ============================================================== #
# Calendario personal del estudiante
# ============================================================== #

class MyCalendarView(LoginRequiredMixin, TemplateView):
    template_name = "activities/my_calendar.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        enrollments = Enrollment.objects.filter(user=self.request.user).select_related("activity", "schedule")
        calendar_dict = defaultdict(list)
        for enrollment in enrollments:
            activity = enrollment.activity
            schedule = enrollment.schedule
            if schedule:
                calendar_dict[schedule.day].append({
                    "name": activity.name,
                    "category": activity.category,
                    "type": activity.type,
                    "location": activity.location,
                    "start_time": schedule.start_time,
                    "end_time": schedule.end_time,
                    "requires_registration": activity.requires_registration,
                    "enrollment_id": enrollment.id,
                })

        day_order = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        sorted_calendar = {day: calendar_dict[day] for day in day_order}
        context["calendar"] = sorted_calendar
        return context


class UnenrollFromActivityView(LoginRequiredMixin, View):
    def post(self, request, enrollment_id):
        enrollment = get_object_or_404(Enrollment, pk=enrollment_id, user=request.user)
        activity_name = enrollment.activity.name
        enrollment.delete()
        messages.success(request, f"La actividad '{activity_name}' ha sido eliminada de tu calendario.")
        return redirect("my_calendar")

    def get(self, request, enrollment_id):
        messages.warning(request, "Método no permitido para esta acción.")
        return redirect("my_calendar")


class UnifiedCalendarView(LoginRequiredMixin, TemplateView):
    template_name = "activities/unified_calendar.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # === Determinar mes y año actuales o solicitados ===
        today = timezone.localtime().date()
        year = int(self.request.GET.get("year", today.year))
        month = int(self.request.GET.get("month", today.month))
        user = self.request.user

        cal = calendar.Calendar(firstweekday=0)  # lunes = 0
        month_days = cal.monthdatescalendar(year, month)

        actividades_normales = Activity.objects.filter(
            is_published=True
        ).exclude(type="Eventos").prefetch_related("schedules", "enrollments")
        
        actividades_eventos = Activity.objects.filter(
            is_published=True,
            type="Eventos"
        ).prefetch_related("schedules")
        
        torneos = Tournament.objects.filter(
            start_date__year=year,
            start_date__month=month
        )
        
        # Partidos de torneos en el mes/año actual
        tournament_games = TournamentGame.objects.filter(
            schedule__date__year=year,
            schedule__date__month=month
        ).select_related('tournament', 'schedule', 'home_team', 'guest_team', 'home_player', 'guest_player')
        
        # Eventos sociales con fecha en el mes/año actual
        eventos_sociales = SocialEvent.objects.filter(
            event_date__year=year,
            event_date__month=month
        ).select_related("project")

        # Obtener inscripciones del usuario para actividades normales
        user_enrollments = set(
            Enrollment.objects.filter(user=user).values_list('activity_id', flat=True)
        )
        
        # Obtener participaciones existentes del usuario para actividades
        user_participations_activities = set(
            Participation.objects.filter(
                user=user,
                attendance_date__year=year,
                attendance_date__month=month
            ).values_list('activity_id', 'attendance_date')
        )
        
        # Para eventos sociales, usar SocialEventEnrollment como participación
        # (los eventos sociales son abiertos, así que el enrollment es la participación)
        user_participations_social = set(
            SocialEventEnrollment.objects.filter(
                user=user,
                event__event_date__year=year,
                event__event_date__month=month
            ).values_list('event_id', 'event__event_date')
        )

        def extract_event_date(description):
            """Extrae la fecha del formato FECHA:YYYY-MM-DD HH:MM"""
            import re
            if not description:
                return None
            match = re.search(r'FECHA:(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2})', description)
            if match:
                try:
                    fecha_str = match.group(1)
                    hora_str = match.group(2)
                    return datetime.strptime(f"{fecha_str} {hora_str}", "%Y-%m-%d %H:%M")
                except ValueError:
                    return None
            return None

        # === Función helper para extraer hora de description de eventos sociales ===
        def extract_social_event_time(description):
            """Extrae la hora del formato 'Horario: 10:00 AM' o 'Horario: 2:00 PM'"""
            import re
            if not description:
                return None
            # Buscar patrones como "Horario: 10:00 AM" o "Horario: 2:00 PM" o "Horario: 10:00 a.m."
            match = re.search(r'Horario:\s*(\d{1,2}):(\d{2})\s*(AM|PM|a\.m\.|p\.m\.)', description, re.IGNORECASE)
            if match:
                try:
                    hour = int(match.group(1))
                    minute = int(match.group(2))
                    am_pm = match.group(3).upper().replace('.', '')  # Remover puntos para comparación
                    
                    # Convertir a formato 24 horas
                    if 'PM' in am_pm:
                        if hour != 12:
                            hour += 12
                    elif hour == 12:  # 12 AM = medianoche
                        hour = 0
                    
                    return f"{hour:02d}:{minute:02d}"
                except (ValueError, IndexError):
                    return None
            return None

        # === Mapeo de días de la semana (reutilizable) ===
        weekday_map = {
            0: "Lunes",
            1: "Martes",
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sábado",
            6: "Domingo",
        }

        # === Estructura base del calendario ===
        calendar_days = []
        for week in month_days:
            for day in week:
                day_data = {
                    "day": day.day,
                    "date": day,
                    "current_month": (day.month == month),
                    "items": []
                }

                for evento_act in actividades_eventos:
                    evento_fecha = extract_event_date(evento_act.description)
                    if evento_fecha and evento_fecha.date() == day:
                        hora_24 = evento_fecha.time()
                        if hora_24.hour == 0:
                            hora_str = f"12:{hora_24.minute:02d} a.m."
                        elif hora_24.hour < 12:
                            hora_str = f"{hora_24.hour}:{hora_24.minute:02d} a.m."
                        elif hora_24.hour == 12:
                            hora_str = f"12:{hora_24.minute:02d} p.m."
                        else:
                            hora_str = f"{hora_24.hour - 12}:{hora_24.minute:02d} p.m."
                        
                        dia_semana = weekday_map.get(evento_fecha.weekday(), "")
                        event_date = evento_fecha.date()
                        
                        # Los eventos institucionales son abiertos - todos pueden registrar participación
                        # Verificar si ya registró participación en este día
                        has_participated = (evento_act.activityId, event_date) in user_participations_activities
                        
                        # Solo permitir registro el día actual
                        can_register = not has_participated and event_date == today
                        
                        day_data["items"].append({
                            "id": evento_act.activityId,
                            "item_type": "event",
                            "type": "Evento Institucional",
                            "title": evento_act.name,
                            "fecha": evento_fecha.strftime("%Y-%m-%d"),
                            "dia_semana": dia_semana,
                            "time": hora_str,
                            "location": evento_act.location,
                            "color": "#6d28d9",
                            "can_register": can_register,
                            "has_participated": has_participated,
                            "requires_enrollment": False  # Eventos son abiertos
                        })
                
                for torneo in torneos:
                    if torneo.start_date == day:
                        hora_str = "9:00 a.m."
                        
                        dia_semana = weekday_map.get(torneo.start_date.weekday(), "")
                        
                        location_map = {
                            "Tenis de Mesa": "Coliseo 2",
                            "Baloncesto": "Coliseo 1",
                            "Fútbol": "Canchas",
                            "Voleibol": "Coliseo",
                        }
                        location = location_map.get(torneo.sport, "Coliseo")
                        
                        day_data["items"].append({
                            "id": f"torneo-{torneo.id}",
                            "type": "Evento Institucional",
                            "title": f"Inicia: {torneo.name}",
                            "fecha": torneo.start_date.strftime("%Y-%m-%d"),
                            "dia_semana": dia_semana,
                            "time": hora_str,
                            "location": location,
                            "color": "#6d28d9"
                        })

                # Agregar partidos de torneos
                for game in tournament_games:
                    if game.schedule.date == day:
                        # Formatear hora en formato 12h
                        game_time = game.schedule.start_time
                        if game_time.hour == 0:
                            hora_str = f"12:{game_time.minute:02d} a.m."
                        elif game_time.hour < 12:
                            hora_str = f"{game_time.hour}:{game_time.minute:02d} a.m."
                        elif game_time.hour == 12:
                            hora_str = f"12:{game_time.minute:02d} p.m."
                        else:
                            hora_str = f"{game_time.hour - 12}:{game_time.minute:02d} p.m."
                        
                        # Obtener nombres de equipos/jugadores
                        if game.tournament.modality == 'I':
                            game_title = f"{game.home_player.name if game.home_player else 'TBD'} vs {game.guest_player.name if game.guest_player else 'TBD'}"
                        else:
                            game_title = f"{game.home_team.name if game.home_team else 'TBD'} vs {game.guest_team.name if game.guest_team else 'TBD'}"
                        
                        dia_semana = weekday_map.get(game.schedule.date.weekday(), "")
                        
                        day_data["items"].append({
                            "id": f"game-{game.id}",
                            "type": "Partido",
                            "title": game_title,
                            "fecha": game.schedule.date.strftime("%Y-%m-%d"),
                            "dia_semana": dia_semana,
                            "time": hora_str,
                            "location": game.schedule.space if game.schedule else "Por confirmar",
                            "color": "#ef4444"  # Color rojo para partidos
                        })
                
                # === Procesar eventos sociales ===
                for evento_social in eventos_sociales:
                    if evento_social.event_date == day:
                        # Intentar extraer la hora de la descripción
                        hora_extracted = extract_social_event_time(evento_social.description)
                        
                        if hora_extracted:
                            # Convertir hora 24h a formato 12h con a.m./p.m.
                            try:
                                hora_obj = datetime.strptime(hora_extracted, "%H:%M").time()
                                if hora_obj.hour == 0:
                                    hora_str = f"12:{hora_obj.minute:02d} a.m."
                                elif hora_obj.hour < 12:
                                    hora_str = f"{hora_obj.hour}:{hora_obj.minute:02d} a.m."
                                elif hora_obj.hour == 12:
                                    hora_str = f"12:{hora_obj.minute:02d} p.m."
                                else:
                                    hora_str = f"{hora_obj.hour - 12}:{hora_obj.minute:02d} p.m."
                            except ValueError:
                                hora_str = "Horario por confirmar"
                        else:
                            # Si no hay hora en la descripción, usar hora por defecto
                            hora_str = "10:00 a.m."
                        
                        # Obtener nombre del día en español
                        dia_semana = weekday_map.get(evento_social.event_date.weekday(), "")
                        
                        # Los eventos sociales son abiertos - todos pueden registrar participación
                        # Verificar si ya registró participación (usando SocialEventEnrollment)
                        # user_participations_social contiene (event_id, event_date)
                        has_participated = (evento_social.id, evento_social.event_date) in user_participations_social
                        
                        # Solo permitir registro el día actual
                        can_register = not has_participated and evento_social.event_date == today
                        
                        day_data["items"].append({
                            "id": f"social-event-{evento_social.id}",
                            "item_type": "social_event",
                            "type": "Proyecto Social",
                            "title": evento_social.name,
                            "fecha": evento_social.event_date.strftime("%Y-%m-%d"),
                            "dia_semana": dia_semana,
                            "time": hora_str,
                            "location": evento_social.location,
                            "color": "#10b981",  # Color verde para proyectos sociales
                            "can_register": can_register,
                            "has_participated": has_participated,
                            "requires_enrollment": False  # Eventos sociales son abiertos
                        })

                # Mapear día de la semana en inglés a español
                weekday_name = calendar.day_name[day.weekday()]  
                weekday_map_english_to_spanish = {
                    "Monday": "Lunes",
                    "Tuesday": "Martes",
                    "Wednesday": "Miércoles",
                    "Thursday": "Jueves",
                    "Friday": "Viernes",
                    "Saturday": "Sábado",
                    "Sunday": "Domingo",
                }
                weekday_spanish = weekday_map_english_to_spanish[weekday_name]

                for act in actividades_normales:
                    for sched in act.schedules.all():
                        if sched.day == weekday_spanish:
                            # Verificar si el usuario está inscrito (si la actividad lo requiere)
                            is_enrolled = act.activityId in user_enrollments
                            # Verificar si ya registró participación en este día
                            has_participated = (act.activityId, day) in user_participations_activities
                            
                            # Determinar si puede registrar participación (solo para el día actual)
                            can_register = False
                            if day == today:  # Solo permitir registro el día actual
                                if act.requires_registration:
                                    # Si requiere inscripción, solo puede registrar si está inscrito
                                    can_register = is_enrolled and not has_participated
                                else:
                                    # Si no requiere inscripción, todos pueden registrar
                                    can_register = not has_participated
                            
                            day_data["items"].append({
                                "id": act.activityId,
                                "item_type": "activity",
                                "type": "Actividad",
                                "title": act.name,
                                "time": sched.start_time.strftime("%H:%M"),
                                "location": act.location,
                                "color": "#8b5cf6",
                                "activity_id": act.activityId,                  
                                "fecha": day.strftime("%Y-%m-%d"),              
                                "hora": sched.start_time.strftime("%H:%M:%S"),  
                                "can_register": can_register,
                                "has_participated": has_participated,
                                "requires_enrollment": act.requires_registration,
                                "is_enrolled": is_enrolled
                        })

                calendar_days.append(day_data)

        # === Navegación de mes ===
        prev_month = month - 1 or 12
        next_month = month + 1 if month < 12 else 1
        prev_year = year - 1 if month == 1 else year
        next_year = year + 1 if month == 12 else year
        month_name = calendar.month_name[month]

        context["calendar_days"] = calendar_days
        context["month_name"] = month_name
        context["year"] = year
        context["prev_month"] = prev_month
        context["next_month"] = next_month
        context["prev_year"] = prev_year
        context["next_year"] = next_year
        context["today"] = today
        context["today_str"] = today.strftime("%Y-%m-%d")

        return context


class ParticipationSegmentationView(LoginRequiredMixin, TemplateView):
    template_name = "activities/cadi_segmentation_participation.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        activity_type = self.request.GET.get("activity_type")
        
        admin_group = Group.objects.filter(name='admin').first()
        admin_user_ids = set()
        if admin_group:
            admin_user_ids.update(admin_group.user_set.values_list('id', flat=True))
        admin_user_ids.update(
            CustomUser.objects.filter(
                Q(is_staff=True) | Q(is_superuser=True) | Q(faculty__name='CADI')
            ).values_list('id', flat=True)
        )
        
        # === ACTIVIDADES NORMALES Y EVENTOS INSTITUCIONALES ===
        enrollments = Enrollment.objects.exclude(user__id__in=admin_user_ids).select_related(
            "activity", "schedule", "user"
        )
        
        participations = Participation.objects.filter(
            attendance_date__isnull=False
        ).exclude(user__id__in=admin_user_ids).select_related(
            "activity", "schedule", "user"
        )

        # === OBTENER PARÁMETROS DE FILTRO ===
        period_filter = self.request.GET.get("period_filter")  # 'semanal', 'mensual', None
        user_type_filter = self.request.GET.get("user_type")  # 'estudiante', 'profesor', 'egresado', 'externo', None
        schedule_filter = self.request.GET.get("schedule")  # Horario específico (ej: "Lunes 10:00-12:00")
        
        # Filtrar por tipo de actividad (si no es "Proyecto Social")
        show_social_events = True
        if activity_type:
            if activity_type == "Proyecto Social":
                # Si se selecciona "Proyecto Social", solo mostrar eventos sociales
                enrollments = enrollments.none()  # No mostrar actividades
                participations = participations.none()  # No mostrar participaciones de actividades
                show_social_events = True
            else:
                # Filtrar actividades normales por tipo
                enrollments = enrollments.filter(activity__type=activity_type)
                participations = participations.filter(activity__type=activity_type)
                show_social_events = False  # No mostrar eventos sociales
        
        # === FILTRAR POR TIPO DE USUARIO ===
        # Usar grupos de Django para identificar tipos de usuario
        if user_type_filter:
            # Mapear tipos de usuario a grupos (si existen)
            group_mapping = {
                'estudiante': 'basic user',  # Asumiendo que estudiantes están en "basic user"
                'profesor': 'profesor',  # Si existe grupo de profesores
                'egresado': 'egresado',  # Si existe grupo de egresados
                'externo': 'externo',  # Si existe grupo de externos
            }
            group_name = group_mapping.get(user_type_filter)
            if group_name:
                group = Group.objects.filter(name=group_name).first()
                if group:
                    user_ids_in_group = set(group.user_set.values_list('id', flat=True))
                    enrollments = enrollments.filter(user__id__in=user_ids_in_group)
                    participations = participations.filter(user__id__in=user_ids_in_group)
        
        # === FILTRAR POR HORARIO ===
        if schedule_filter:
            # schedule_filter viene como "Lunes 10:00-12:00" o similar
            # Extraer día y rango de horas
            schedule_parts = schedule_filter.split()
            if len(schedule_parts) >= 2:
                day = schedule_parts[0]
                time_range = ' '.join(schedule_parts[1:])  # "10:00-12:00"
                if '-' in time_range:
                    start_time_str, end_time_str = time_range.split('-')
                    try:
                        start_time = datetime.strptime(start_time_str.strip(), "%H:%M").time()
                        end_time = datetime.strptime(end_time_str.strip(), "%H:%M").time()
                        # Filtrar enrollments y participations por schedule
                        enrollments = enrollments.filter(
                            schedule__day=day,
                            schedule__start_time=start_time,
                            schedule__end_time=end_time
                        )
                        participations = participations.filter(
                            schedule__day=day,
                            schedule__start_time=start_time,
                            schedule__end_time=end_time
                        )
                    except ValueError:
                        pass  # Si no se puede parsear, ignorar el filtro

        # === CONSTRUIR activity_groups: Agrupar por actividad y horario ===
        activity_groups = defaultdict(lambda: {
            "activity": None,
            "schedule_info": None,
            "enrolled_students": [],
            "participated_students": defaultdict(list),  # Key: attendance_date, Value: list of students
            "total_enrolled": 0,
            "total_participated_by_date": defaultdict(int),
        })

        # Procesar enrollments (inscripciones)
        for enrollment in enrollments:
            activity = enrollment.activity
            
            if not activity:
                continue
            
            schedule = enrollment.schedule
            
            if schedule:
                schedule_key = f"{schedule.day} {schedule.start_time}-{schedule.end_time}"
            else:
                schedule_key = "Sin horario"
            
            key = (activity.activityId, schedule_key)
            
            student_id = enrollment.user.identification or "N/A"
            student_info = {
                "id": student_id,
                "name": enrollment.user.get_full_name() or enrollment.user.username,
                "username": enrollment.user.username,
                "type": "enrolled",
            }
            
            existing_enrolled_ids = {s["id"] for s in activity_groups[key]["enrolled_students"]}
            if student_id not in existing_enrolled_ids:
                activity_groups[key]["enrolled_students"].append(student_info)
            
            activity_groups[key]["activity"] = activity
            activity_groups[key]["schedule_info"] = schedule_key
            activity_groups[key]["total_enrolled"] = len(activity_groups[key]["enrolled_students"])

        # Procesar participations (asistencias reales)
        for participation in participations:
            activity = participation.activity
            attendance_date = participation.attendance_date
            schedule = participation.schedule
            
            if not activity or not attendance_date:
                continue
            
            if schedule:
                schedule_key = f"{schedule.day} {schedule.start_time}-{schedule.end_time}"
            else:
                schedule_key = "Sin horario"
            
            key = (activity.activityId, schedule_key)
            
            student_id = participation.user.identification or "N/A"
            student_info = {
                "id": student_id,
                "name": participation.user.get_full_name() or participation.user.username,
                "username": participation.user.username,
                "attendance_date": attendance_date,
                "attendance_time": participation.attendance_time,
                "type": "participated",
            }
            
            if activity_groups[key]["activity"] is None:
                activity_groups[key]["activity"] = activity
                activity_groups[key]["schedule_info"] = schedule_key
            
            date_key = attendance_date.isoformat()
            existing_participated_ids = {s["id"] for s in activity_groups[key]["participated_students"][date_key]}
            if student_id not in existing_participated_ids:
                activity_groups[key]["participated_students"][date_key].append(student_info)
                activity_groups[key]["total_participated_by_date"][date_key] = len(activity_groups[key]["participated_students"][date_key])

        # Transformar a lista para el template
        grouped_data = []
        for (activity_id, schedule_key), info in activity_groups.items():
            activity = info["activity"]
            
            if activity is None:
                continue
            
            info["enrolled_students"].sort(key=lambda x: x["id"])
            
            if info["participated_students"]:
                participation_dates = sorted(info["participated_students"].keys())
                
                for date_key in participation_dates:
                    students = info["participated_students"][date_key]
                    students.sort(key=lambda x: x["id"])
                    attendance_date = datetime.strptime(date_key, "%Y-%m-%d").date()
                    
                    grouped_data.append({
                        "period": attendance_date,
                        "activity_name": activity.name,
                        "activity_type": activity.type or "N/A",
                        "activity_id": activity.activityId,
                        "schedule": schedule_key,
                        "total_enrolled": info["total_enrolled"],
                        "total_participated": len(students),
                        "enrolled_students": info["enrolled_students"],
                        "participated_students": students,
                        "has_participations": True,
                        "is_social_event": False,
                    })
            else:
                if info["total_enrolled"] > 0:
                    grouped_data.append({
                        "period": None,
                        "activity_name": activity.name,
                        "activity_type": activity.type or "N/A",
                        "activity_id": activity.activityId,
                        "schedule": schedule_key,
                        "total_enrolled": info["total_enrolled"],
                        "total_participated": 0,
                        "enrolled_students": info["enrolled_students"],
                        "participated_students": [],
                        "has_participations": False,
                        "is_social_event": False,
                    })

        # === EVENTOS SOCIALES ===
        if show_social_events:
            # Obtener todas las inscripciones (todos los que se inscribieron)
            social_event_enrollments_all = SocialEventEnrollment.objects.exclude(
                user__id__in=admin_user_ids
            ).select_related("event", "user", "event__project")
            
            # Aplicar filtro por tipo de usuario a eventos sociales
            if user_type_filter:
                group_mapping = {
                    'estudiante': 'basic user',
                    'profesor': 'profesor',
                    'egresado': 'egresado',
                    'externo': 'externo',
                }
                group_name = group_mapping.get(user_type_filter)
                if group_name:
                    group = Group.objects.filter(name=group_name).first()
                    if group:
                        user_ids_in_group = set(group.user_set.values_list('id', flat=True))
                        social_event_enrollments_all = social_event_enrollments_all.filter(user__id__in=user_ids_in_group)
            
            # Agrupar por evento y fecha del evento
            social_event_groups = defaultdict(lambda: {
                "event": None,
                "enrolled_students": [],
                "participated_students": [],
                "total_enrolled": 0,
                "total_participated": 0,
            })
            
            for enrollment in social_event_enrollments_all:
                event = enrollment.event
                event_date = event.event_date
                
                key = (event.id, event_date)
                
                student_id = enrollment.user.identification or "N/A"
                student_name = enrollment.user.get_full_name() or enrollment.user.username
                
                # Todos los enrollments son inscritos
                enrolled_student_info = {
                    "id": student_id,
                    "name": student_name,
                    "username": enrollment.user.username,
                    "type": "enrolled",
                }
                
                social_event_groups[key]["event"] = event
                
                # Agregar a inscritos si no existe
                existing_enrolled_ids = {s["id"] for s in social_event_groups[key]["enrolled_students"]}
                if student_id not in existing_enrolled_ids:
                    social_event_groups[key]["enrolled_students"].append(enrolled_student_info)
                    social_event_groups[key]["total_enrolled"] = len(social_event_groups[key]["enrolled_students"])
                
                is_participant = False
                if enrollment.confirmed and enrollment.confirmed_at:
                    # Calcular diferencia de tiempo entre confirmed_at y enrolled_at
                    time_diff = (enrollment.confirmed_at - enrollment.enrolled_at).total_seconds()
                    
                    # Calcular fecha de creación del enrollment
                    enrollment_date = enrollment.enrolled_at.date()
                    today = timezone.localtime().date()
                    
                    if time_diff > 0.05:
                        is_participant = True
                    elif time_diff <= 0.05:
                        # Si se creó en el día del evento o después, es probable que sea desde el calendario
                        if enrollment_date >= event_date:
                            is_participant = True
                        # Si se creó antes del evento Y el evento ya pasó, también considerar como participante
                        # (porque los eventos sociales son abiertos)
                        elif enrollment_date < event_date and event_date <= today:
                            # El evento ya pasó y había un enrollment, probablemente participó
                            is_participant = True
                        # Si se creó antes del evento Y el evento aún no ha pasado, es solo inscripción
                        else:
                            is_participant = False
                
                if is_participant:
                    participated_student_info = {
                        "id": student_id,
                        "name": student_name,
                        "username": enrollment.user.username,
                        "attendance_date": event_date,
                        "attendance_time": None,  # Los eventos sociales no tienen hora específica de asistencia
                        "type": "participated",
                    }
                    
                    existing_participated_ids = {s["id"] for s in social_event_groups[key]["participated_students"]}
                    if student_id not in existing_participated_ids:
                        social_event_groups[key]["participated_students"].append(participated_student_info)
                        social_event_groups[key]["total_participated"] = len(social_event_groups[key]["participated_students"])
            
            # Agregar eventos sociales a grouped_data
            for (event_id, event_date), info in social_event_groups.items():
                event = info["event"]
                
                if event is None:
                    continue
                
                # Ordenar estudiantes por ID
                info["enrolled_students"].sort(key=lambda x: x["id"])
                info["participated_students"].sort(key=lambda x: x["id"])
                
                # Extraer hora de la descripción si existe
                hora_str = "Por confirmar"
                if event.description:
                    match = re.search(r'Horario:\s*(\d{1,2}):(\d{2})\s*(AM|PM|a\.m\.|p\.m\.)', event.description, re.IGNORECASE)
                    if match:
                        try:
                            hour = int(match.group(1))
                            minute = int(match.group(2))
                            am_pm = match.group(3).upper().replace('.', '')
                            if 'PM' in am_pm:
                                if hour != 12:
                                    hour += 12
                            elif hour == 12:
                                hour = 0
                            hora_str = f"{hour:02d}:{minute:02d}"
                        except (ValueError, IndexError):
                            pass
                
                grouped_data.append({
                    "period": event_date,
                    "activity_name": event.name,
                    "activity_type": "Proyecto Social",
                    "activity_id": f"social-event-{event.id}",
                    "schedule": hora_str if hora_str != "Por confirmar" else "Evento social",
                    "location": event.location,
                    "total_enrolled": info["total_enrolled"],  # Todos los que se inscribieron
                    "total_participated": info["total_participated"],  # Solo los que registraron participación
                    "enrolled_students": info["enrolled_students"],  # Lista de inscritos
                    "participated_students": info["participated_students"],  # Lista de participantes
                    "has_participations": info["total_participated"] > 0,
                    "is_social_event": True,
                })
        
        # === FILTRAR POR PERÍODO (semanal/mensual) ===
        if period_filter:
            today = timezone.localtime().date()
            filtered_data = []
            
            for item in grouped_data:
                if not item["period"]:
                    # Si no tiene período (solo inscripciones), no incluirlo en filtros de período
                    continue
                
                item_date = item["period"]
                
                if period_filter == "semanal":
                    # Filtrar por semana actual (últimos 7 días)
                    week_start = today - timedelta(days=today.weekday())
                    week_end = week_start + timedelta(days=6)
                    if week_start <= item_date <= week_end:
                        filtered_data.append(item)
                elif period_filter == "mensual":
                    # Filtrar por mes actual
                    if item_date.year == today.year and item_date.month == today.month:
                        filtered_data.append(item)
            
            grouped_data = filtered_data
        
        grouped_data.sort(key=lambda x: (
            x["activity_name"], 
            x["period"] if x["period"] else datetime.max.date()
        ))
        
        # === OBTENER HORARIOS DISPONIBLES PARA EL FILTRO ===
        # Obtener todos los horarios únicos de las actividades
        all_schedules = Schedule.objects.select_related("activity").distinct()
        schedule_choices = []
        seen_schedules = set()
        for schedule in all_schedules:
            # Formatear tiempo sin segundos (solo HH:MM) para que coincida con el filtro
            start_time_str = schedule.start_time.strftime("%H:%M")
            end_time_str = schedule.end_time.strftime("%H:%M")
            schedule_str = f"{schedule.day} {start_time_str}-{end_time_str}"
            if schedule_str not in seen_schedules:
                schedule_choices.append((schedule_str, schedule_str))
                seen_schedules.add(schedule_str)
        schedule_choices.sort()

        activity_choices = Activity._meta.get_field("type").choices
        # Agregar "Proyecto Social" a las opciones de filtro
        activity_choices_list = list(activity_choices)
        activity_choices_list.append(("Proyecto Social", "Proyecto Social"))

        context.update({
            "grouped_data": grouped_data,
            "activity_choices": activity_choices_list,
            "selected_type": activity_type,
            "selected_period": period_filter,
            "selected_user_type": user_type_filter,
            "selected_schedule": schedule_filter,
            "schedule_choices": schedule_choices,
            "user_type_choices": [
                ("estudiante", "Estudiante"),
                ("profesor", "Profesor"),
                ("egresado", "Egresado"),
                ("externo", "Externo"),
            ],
            "period_choices": [
                ("semanal", "Semanal"),
                ("mensual", "Mensual"),
            ],
        })
        
        return context


def _get_segmentation_data(activity_type=None):
    enrollments = Enrollment.objects.select_related("activity", "schedule", "user")
    participations = Participation.objects.select_related("activity", "schedule", "user")

    if activity_type:
        enrollments = enrollments.filter(activity__type=activity_type)
        participations = participations.filter(activity__type=activity_type)

    enrollments_grouped = (
        enrollments
        .annotate(period=TruncDate("registered_at"))
        .values(
            "period",
            "activity__name",
            "activity__type",
            "activity__activityId",
            "schedule__day",
            "schedule__start_time",
            "schedule__end_time",
        )
        .annotate(total_participants=Count("user", distinct=True))
        .order_by("-period")
    )

    participations_grouped = (
        participations
        .annotate(period=TruncDate("attendance_date"))
        .values(
            "period",
            "activity__name",
            "activity__type",
            "activity__activityId",
            "schedule__day",
            "schedule__start_time",
            "schedule__end_time",
        )
        .annotate(total_participants=Count("user", distinct=True))
        .order_by("-period")
    )

    merged = defaultdict(lambda: {"total_participants": 0})

    for row in list(enrollments_grouped) + list(participations_grouped):
        key = (
            row["period"],
            row["activity__name"],
            row["activity__type"],
            row["schedule__day"],
            row["schedule__start_time"],
            row["schedule__end_time"],
        )
        merged[key]["total_participants"] += row["total_participants"]

    grouped_data = []
    for (period, name, type_, day, start, end), info in merged.items():
        period_str = period.strftime("%Y-%m-%d") if period and hasattr(period, 'strftime') else (str(period) if period else "-")
        grouped_data.append({
            "period": period_str,
            "activity_name": name or "-",
            "activity_type": type_ or "-",
            "day": day or "-",
            "schedule": f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}" if start and end and hasattr(start, 'strftime') and hasattr(end, 'strftime') else "-",
            "total_participants": info["total_participants"],
        })

    grouped_data.sort(key=lambda x: (x["activity_name"], x["day"]))
    return grouped_data


class DownloadSegmentationExcelView(LoginRequiredMixin, View):
    def get(self, request):
        activity_type = request.GET.get("activity_type")
        data = _get_segmentation_data(activity_type)

        wb = Workbook()
        ws = wb.active
        ws.title = "Segmentación de Participación"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Periodo", "Actividad", "Tipo", "Día", "Horario", "Total de Participantes"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        for row_num, row_data in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=row_data["period"])
            ws.cell(row=row_num, column=2, value=row_data["activity_name"])
            ws.cell(row=row_num, column=3, value=row_data["activity_type"])
            ws.cell(row=row_num, column=4, value=row_data["day"])
            ws.cell(row=row_num, column=5, value=row_data["schedule"])
            ws.cell(row=row_num, column=6, value=row_data["total_participants"])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="segmentacion_participacion.xlsx"'
        wb.save(response)
        return response


class DownloadSegmentationCSVView(LoginRequiredMixin, View):
    def get(self, request):
        activity_type = request.GET.get("activity_type")
        data = _get_segmentation_data(activity_type)

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="segmentacion_participacion.csv"'
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(["Periodo", "Actividad", "Tipo", "Día", "Horario", "Total de Participantes"])

        for row_data in data:
            writer.writerow([
                row_data["period"],
                row_data["activity_name"],
                row_data["activity_type"],
                row_data["day"],
                row_data["schedule"],
                row_data["total_participants"],
            ])

        return response


class RegisterParticipationView(LoginRequiredMixin, View):
    """
    Permite a los estudiantes registrar su propia asistencia a:
    - Actividades (requiere inscripción si requires_registration=True)
    - Eventos institucionales (abiertos a todos)
    - Proyectos sociales (abiertos a todos)
    """
    def post(self, request):
        item_type = request.POST.get('item_type')  # 'activity', 'event', 'social_event'
        item_id = request.POST.get('item_id')  # activity_id, event_id, social_event_id
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')
        user = request.user

        if not item_type or not item_id:
            messages.error(request, "❌ No se recibió la información necesaria.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Parsear fecha
        if fecha:
            try:
                attendance_date = datetime.strptime(fecha, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "❌ Formato de fecha inválido. Use YYYY-MM-DD.")
                return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            # If no fecha provided, use today
            attendance_date = timezone.localtime().date()

        # Parsear hora - puede venir en formato "10:00 a.m." o "14:00" o "14:00:00"
        attendance_time = None
        if hora:
            # Try parsing hora in different formats
            try:
                # Limpiar formato: remover " a.m." y " p.m." y espacios
                hora_clean = hora.replace(' a.m.', '').replace(' p.m.', '').replace(' AM', '').replace(' PM', '').strip()
                
                if ':' in hora_clean:
                    parts = hora_clean.split(':')
                    if len(parts) == 2:
                        hour = int(parts[0])
                        minute = int(parts[1])
                        # Verificar si era PM (buscar en el string original)
                        if 'p.m.' in hora.lower() or 'pm' in hora.lower():
                            if hour != 12:
                                hour += 12
                        elif 'a.m.' in hora.lower() or 'am' in hora.lower():
                            if hour == 12:
                                hour = 0
                        attendance_time = datetime.strptime(f"{hour:02d}:{minute:02d}", "%H:%M").time()
                    elif len(parts) == 3:
                        hour = int(parts[0])
                        minute = int(parts[1])
                        # Verificar si era PM
                        if 'p.m.' in hora.lower() or 'pm' in hora.lower():
                            if hour != 12:
                                hour += 12
                        elif 'a.m.' in hora.lower() or 'am' in hora.lower():
                            if hour == 12:
                                hour = 0
                        attendance_time = datetime.strptime(f"{hour:02d}:{minute:02d}:{parts[2]}", "%H:%M:%S").time()
            except (ValueError, IndexError, AttributeError):
                attendance_time = None

        # === MANEJAR ACTIVIDADES NORMALES ===
        if item_type == 'activity':
            try:
                activity = get_object_or_404(Activity, activityId=int(item_id))
            except ValueError:
                messages.error(request, "❌ ID de actividad inválido.")
                return redirect(request.META.get('HTTP_REFERER', '/'))

            # Validar inscripción si la actividad la requiere
            if activity.requires_registration:
                enrollment = Enrollment.objects.filter(
                    user=user,
                    activity=activity
                ).first()
                
                if not enrollment:
                    messages.error(
                        request,
                        f"❌ Debes estar inscrito en '{activity.name}' para registrar tu participación."
                    )
                    return redirect(request.META.get('HTTP_REFERER', '/'))

            # Buscar schedule
            schedule = None
            if hora and attendance_time:
                schedule = activity.schedules.filter(
                    start_time__lte=attendance_time,
                    end_time__gte=attendance_time
                ).first()

            if not schedule:
                weekday_map = {
                    0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves",
                    4: "Viernes", 5: "Sábado", 6: "Domingo"
                }
                weekday_spanish = weekday_map[attendance_date.weekday()]
                schedule = activity.schedules.filter(day=weekday_spanish).first()

            # Verificar si ya existe participación
            existing = Participation.objects.filter(
                activity=activity,
                user=user,
                attendance_date=attendance_date
            ).first()

            if existing:
                messages.warning(
                    request,
                    f"⚠️ Ya registraste tu participación en '{activity.name}' el {attendance_date.strftime('%d/%m/%Y')}."
                )
                return redirect(request.META.get('HTTP_REFERER', '/'))

            # Crear participación
            Participation.objects.create(
                activity=activity,
                user=user,
                attendance_date=attendance_date,
                attendance_time=attendance_time,
                schedule=schedule
            )

            messages.success(
                request,
                f"✅ Participación registrada para '{activity.name}' el {attendance_date.strftime('%d/%m/%Y')}."
            )

        # === MANEJAR EVENTOS INSTITUCIONALES (type="Eventos") ===
        elif item_type == 'event':
            try:
                activity = get_object_or_404(Activity, activityId=int(item_id), type="Eventos")
            except ValueError:
                messages.error(request, "❌ ID de evento inválido.")
                return redirect(request.META.get('HTTP_REFERER', '/'))

            # Los eventos institucionales son abiertos, no requieren inscripción
            # Verificar si ya existe participación
            existing = Participation.objects.filter(
                activity=activity,
                user=user,
                attendance_date=attendance_date
            ).first()

            if existing:
                messages.warning(
                    request,
                    f"⚠️ Ya registraste tu participación en '{activity.name}' el {attendance_date.strftime('%d/%m/%Y')}."
                )
                return redirect(request.META.get('HTTP_REFERER', '/'))

            # Crear participación (sin schedule para eventos)
            Participation.objects.create(
                activity=activity,
                user=user,
                attendance_date=attendance_date,
                attendance_time=attendance_time,
                schedule=None
            )

            messages.success(
                request,
                f"✅ Participación registrada para '{activity.name}' el {attendance_date.strftime('%d/%m/%Y')}."
            )

        # === MANEJAR EVENTOS SOCIALES ===
        elif item_type == 'social_event':
            try:
                # El item_id viene como "social-event-{id}"
                event_id = int(item_id.replace('social-event-', ''))
                social_event = get_object_or_404(SocialEvent, id=event_id)
            except (ValueError, AttributeError):
                messages.error(request, "❌ ID de evento social inválido.")
                return redirect(request.META.get('HTTP_REFERER', '/'))

            now = timezone.now()
            existing, created = SocialEventEnrollment.objects.get_or_create(
                user=user,
                event=social_event,
                defaults={
                    'confirmed': True,
                }
            )
            
            if not created:
                time.sleep(0.2)  # 200ms de diferencia
                existing.confirmed = True
                existing.confirmed_at = timezone.now()
                existing.save()
                messages.warning(
                    request,
                    f"⚠️ Ya habías registrado tu participación en '{social_event.name}'. Se actualizó la fecha de registro."
                )
            else:
                time.sleep(0.2)  # 200ms de diferencia
                existing.confirmed_at = timezone.now()
                existing.save(update_fields=['confirmed_at'])
                messages.success(
                    request,
                    f"✅ Participación registrada para '{social_event.name}' el {attendance_date.strftime('%d/%m/%Y')}."
                )

        else:
            messages.error(request, "❌ Tipo de item no reconocido.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        return redirect(request.META.get('HTTP_REFERER', '/'))


class ParticipationReportView(LoginRequiredMixin, TemplateView):
    template_name = "activities/participation_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        data = (
            Participation.objects
            .select_related("activity", "user")
            .values("participation_type", "activity__name")
            .annotate(total_participaciones=Count("id"), total_estudiantes=Count("user", distinct=True))
            .order_by("participation_type", "activity__name")
        )
        context["data"] = data
        return context
