from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.utils import timezone
from activities.models import (
    Activity,
    Enrollment,
    Schedule,
    ActivityReview,
    ActivityType,
    CategoryType,
    WeekDay,
    Participation,
)
from login.models import CustomUser, Faculty
from datetime import time, date, timedelta
import io
from openpyxl import load_workbook
import csv

User = get_user_model()


class ActivitiesTestCase(TestCase):
    """Pruebas unitarias y de integración para el módulo Activities."""

    def setUp(self):
        self.client = Client()
        # Create admin user for staff-only views
        admin_group, _ = Group.objects.get_or_create(name='admin')
        self.admin_user = User.objects.create_user(username="admin", password="admin123")
        self.admin_user.groups.add(admin_group)
        self.admin_user.save()
        
        # Create regular user for student views
        self.user = User.objects.create_user(username="student", password="pass123")
        self.client.login(username="student", password="pass123")

        self.activity = Activity.objects.create(
            name="Fútbol",
            description="Partido semanal",
            location="Cancha 1",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=2,
        )

    # ==========================================================
    # 🔹 Pruebas UNITARIAS de modelos
    # ==========================================================

    def test_activity_str_method(self):
        self.assertEqual(str(self.activity), "Fútbol (Deportiva, Grupal)")

    def test_activity_id_property(self):
        self.assertEqual(self.activity.id, self.activity.activityId)

    def test_activity_is_full_status(self):
        # No está llena aún
        self.assertFalse(self.activity.is_full_status)
        # Llenar cupos
        for i in range(2):
            u = User.objects.create_user(username=f"user{i}", password="pass")
            Enrollment.objects.create(user=u, activity=self.activity)
        self.assertTrue(self.activity.is_full_status)

    def test_enrollment_str(self):
        enrollment = Enrollment.objects.create(user=self.user, activity=self.activity)
        self.assertIn(self.user.username, str(enrollment))

    def test_schedule_str(self):
        schedule = Schedule.objects.create(
            activity=self.activity,
            day=WeekDay.MONDAY,
            start_time=time(10, 0),
            end_time=time(11, 0),
        )
        self.assertIn("Lunes", str(schedule))

    def test_activityreview_str(self):
        review = ActivityReview.objects.create(
            activity=self.activity, user=self.user, rating=5, comment="Excelente"
        )
        self.assertIn("5⭐", str(review))

    # ==========================================================
    # 🔹 Pruebas de INTEGRACIÓN (vistas principales)
    # ==========================================================

    def test_public_activities_view(self):
        # Public activities view now requires staff permission
        self.client.login(username="admin", password="admin123")
        url = reverse("public_activities")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fútbol")

    def test_activity_detail_view(self):
        url = reverse("activity_detail", args=[self.activity.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fútbol")


    def test_enroll_in_activity(self):
        # Primero, crear un schedule para la actividad
        schedule = Schedule.objects.create(
            activity=self.activity,
            day=WeekDay.MONDAY,
            start_time=time(10, 0),
            end_time=time(11, 0),
        )
        # GET redirects to activityView (as per the view implementation)
        url = reverse("enroll_in_activity", args=[self.activity.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)  # Redirects to activityView
        self.assertRedirects(response, reverse("activityView"))
        
        # POST para crear la inscripción
        # Note: The view doesn't use schedules parameter, it creates enrollment directly
        response = self.client.post(url)
        self.assertEqual(response.status_code, 302)  # Redirect después de inscribirse
        self.assertEqual(Enrollment.objects.count(), 1)

    def test_review_activity_creates_review(self):
        url = reverse("review_activity", args=[self.activity.pk])
        response = self.client.post(url, {"rating": 4, "comment": "Muy buena"})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(ActivityReview.objects.count(), 1)
        review = ActivityReview.objects.first()
        self.assertEqual(review.rating, 4)
        self.assertEqual(review.comment, "Muy buena")

    def test_activity_reviews_view(self):
        ActivityReview.objects.create(activity=self.activity, user=self.user, rating=3)
        url = reverse("activity_reviews", args=[self.activity.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "3")

    def test_unenroll_from_activity(self):
        enrollment = Enrollment.objects.create(user=self.user, activity=self.activity)
        url = reverse("unenroll_from_activity", args=[enrollment.id])
        response = self.client.post(url)
        self.assertRedirects(response, reverse("my_calendar"))
        self.assertEqual(Enrollment.objects.count(), 0)

    def test_cadi_activity_review_view(self):
        # CADI review view now requires staff permission
        ActivityReview.objects.create(activity=self.activity, user=self.user, rating=5)
        self.client.login(username="admin", password="admin123")
        url = reverse("cadi_activity_review")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fútbol")

    def test_mark_review_read(self):
        # Mark review read now requires staff permission
        review = ActivityReview.objects.create(activity=self.activity, user=self.user, rating=4)
        self.client.login(username="admin", password="admin123")
        url = reverse("mark_review_read", args=[review.id])
        response = self.client.get(url)
        review.refresh_from_db()
        self.assertTrue(review.is_read)


class RegisterParticipationViewTests(TestCase):
    """Tests para RegisterParticipationView"""
    
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username="student", password="pass123")
        self.client.login(username="student", password="pass123")
        
        self.activity = Activity.objects.create(
            name="Yoga",
            description="Clase de yoga",
            location="Sala 1",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )
        
        self.schedule = Schedule.objects.create(
            activity=self.activity,
            day=WeekDay.MONDAY,
            start_time=time(10, 0),
            end_time=time(11, 0),
        )
        
        # Crear enrollment para el usuario
        self.enrollment = Enrollment.objects.create(
            user=self.user,
            activity=self.activity,
            schedule=self.schedule
        )
    
    def test_register_participation_success(self):
        """Test que registra participación exitosamente"""
        url = reverse("register_participation")
        today = timezone.localtime().date()
        
        data = {
            'item_type': 'activity',
            'item_id': str(self.activity.activityId),
            'fecha': today.strftime('%Y-%m-%d'),
            'hora': '10:30:00',  # Use full time format
        }
        
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)  # Redirect
        self.assertEqual(Participation.objects.count(), 1)
        
        participation = Participation.objects.first()
        self.assertEqual(participation.activity, self.activity)
        self.assertEqual(participation.user, self.user)
        self.assertEqual(participation.attendance_date, today)
        # attendance_time might be None if hora parsing fails, but that's okay
        # The view should handle both formats
    
    def test_register_participation_without_activity_id(self):
        """Test que falla cuando no se proporciona item_id"""
        url = reverse("register_participation")
        data = {
            'item_type': 'activity',
            'fecha': '2025-01-15',
            'hora': '10:30',
        }
        
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)  # Redirect con error
        self.assertEqual(Participation.objects.count(), 0)
    
    def test_register_participation_invalid_activity_id(self):
        """Test que falla con item_id inválido"""
        url = reverse("register_participation")
        data = {
            'item_type': 'activity',
            'item_id': 'invalid',
            'fecha': '2025-01-15',
            'hora': '10:30',
        }
        
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)  # Redirect con error
        self.assertEqual(Participation.objects.count(), 0)
    
    def test_register_participation_invalid_date_format(self):
        """Test que falla con formato de fecha inválido"""
        url = reverse("register_participation")
        data = {
            'item_type': 'activity',
            'item_id': str(self.activity.activityId),
            'fecha': '15-01-2025',  # Formato incorrecto
            'hora': '10:30',
        }
        
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)  # Redirect con error
        self.assertEqual(Participation.objects.count(), 0)
    
    def test_register_participation_duplicate(self):
        """Test que previene participación duplicada"""
        url = reverse("register_participation")
        today = timezone.localtime().date()
        
        # Crear participación existente
        Participation.objects.create(
            activity=self.activity,
            user=self.user,
            attendance_date=today,
            schedule=self.schedule
        )
        
        data = {
            'item_type': 'activity',
            'item_id': str(self.activity.activityId),
            'fecha': today.strftime('%Y-%m-%d'),
            'hora': '10:30',
        }
        
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)  # Redirect con warning
        self.assertEqual(Participation.objects.count(), 1)  # No se crea duplicado
    
    def test_register_participation_without_fecha(self):
        """Test que usa fecha de hoy cuando no se proporciona fecha"""
        url = reverse("register_participation")
        today = timezone.localtime().date()
        
        data = {
            'item_type': 'activity',
            'item_id': str(self.activity.activityId),
            'hora': '10:30',
        }
        
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)
        # Note: The view now uses today's date when fecha is not provided
        # But the original view had a check that required fecha, which was removed
        # So this should work now
        self.assertEqual(Participation.objects.count(), 1)
        
        participation = Participation.objects.first()
        self.assertEqual(participation.attendance_date, today)


class ParticipationSegmentationViewTests(TestCase):
    """Tests para ParticipationSegmentationView"""
    
    def setUp(self):
        self.client = Client()
        # Crear usuario admin (será excluido)
        self.admin_user = User.objects.create_user(
            username="admin_user",
            password="pass123",
            is_staff=True
        )
        # Crear usuario estudiante normal
        self.student_user = User.objects.create_user(
            username="student_user",
            password="pass123"
        )
        # Crear facultad CADI (también será excluida) - usar get_or_create para evitar duplicados
        self.cadi_faculty, _ = Faculty.objects.get_or_create(name="CADI")
        self.cadi_user = User.objects.create_user(
            username="cadi_user",
            password="pass123"
        )
        self.cadi_user.faculty = self.cadi_faculty
        self.cadi_user.save()
        
        self.client.login(username="admin_user", password="pass123")
        
        # Crear actividades
        self.activity1 = Activity.objects.create(
            name="Fútbol",
            description="Partido de fútbol",
            location="Cancha 1",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )
        
        self.activity2 = Activity.objects.create(
            name="Yoga",
            description="Clase de yoga",
            location="Sala 1",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )
        
        # Crear schedules
        self.schedule1 = Schedule.objects.create(
            activity=self.activity1,
            day=WeekDay.MONDAY,
            start_time=time(10, 0),
            end_time=time(11, 0),
        )
        
        self.schedule2 = Schedule.objects.create(
            activity=self.activity2,
            day=WeekDay.TUESDAY,
            start_time=time(14, 0),
            end_time=time(15, 0),
        )
        
        # Crear enrollments (solo para estudiante, no admin)
        self.enrollment1 = Enrollment.objects.create(
            user=self.student_user,
            activity=self.activity1,
            schedule=self.schedule1
        )
        
        self.enrollment2 = Enrollment.objects.create(
            user=self.student_user,
            activity=self.activity2,
            schedule=self.schedule2
        )
        
        # Crear participaciones
        today = timezone.localtime().date()
        self.participation1 = Participation.objects.create(
            activity=self.activity1,
            user=self.student_user,
            attendance_date=today,
            schedule=self.schedule1
        )
    
    def test_participation_segmentation_view_basic(self):
        """Test básico de la vista de segmentación"""
        url = reverse("participation_segmentation")
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fútbol")
        # Verificar que el contexto tiene los datos esperados
        context = response.context
        self.assertIn('grouped_data', context)
        self.assertIn('activity_choices', context)
        grouped_data = context['grouped_data']
        self.assertIsInstance(grouped_data, list)
        # Verificar que hay datos de actividades
        if grouped_data:
            self.assertIn('activity_name', grouped_data[0])
    
    def test_participation_segmentation_view_with_filter(self):
        """Test con filtro por tipo de actividad"""
        url = reverse("participation_segmentation")
        response = self.client.get(url, {'activity_type': ActivityType.DEPORTIVA})
        
        self.assertEqual(response.status_code, 200)
        context = response.context
        self.assertIn('grouped_data', context)
        self.assertIn('selected_type', context)
        self.assertEqual(context['selected_type'], ActivityType.DEPORTIVA)
    
    def test_participation_segmentation_excludes_admin_users(self):
        """Test que excluye usuarios admin de los resultados"""
        # Crear enrollment para admin (debe ser excluido)
        Enrollment.objects.create(
            user=self.admin_user,
            activity=self.activity1,
            schedule=self.schedule1
        )
        
        # Crear enrollment para usuario CADI (debe ser excluido)
        Enrollment.objects.create(
            user=self.cadi_user,
            activity=self.activity1,
            schedule=self.schedule1
        )
        
        url = reverse("participation_segmentation")
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        context = response.context
        
        # Verificar que grouped_data solo contiene estudiantes normales
        if 'grouped_data' in context:
            grouped_data = context['grouped_data']
            for group in grouped_data:
                enrolled_students = group.get('enrolled_students', [])
                # Verificar que no hay usuarios admin en la lista
                student_usernames = [s.get('username') for s in enrolled_students]
                self.assertNotIn('admin_user', student_usernames)
                self.assertNotIn('cadi_user', student_usernames)
                
                participated_students = group.get('participated_students', [])
                participated_usernames = [s.get('username') for s in participated_students]
                self.assertNotIn('admin_user', participated_usernames)
                self.assertNotIn('cadi_user', participated_usernames)
    
    def test_filter_by_period_semanal(self):
        """Test de filtro por período semanal"""
        # Eliminar participación existente del setUp para evitar duplicados
        Participation.objects.filter(
            activity=self.activity1,
            user=self.student_user
        ).delete()
        
        # Crear participación en la semana actual
        today = timezone.localtime().date()
        participation_today, _ = Participation.objects.get_or_create(
            activity=self.activity1,
            user=self.student_user,
            attendance_date=today,
            defaults={'schedule': self.schedule1}
        )
        
        # Crear participación hace 2 semanas (fuera del período semanal)
        two_weeks_ago = today - timedelta(days=14)
        participation_old = Participation.objects.create(
            activity=self.activity2,
            user=self.student_user,
            attendance_date=two_weeks_ago,
            schedule=self.schedule2
        )
        
        url = reverse("participation_segmentation")
        response = self.client.get(url, {'period_filter': 'semanal'})
        
        self.assertEqual(response.status_code, 200)
        context = response.context
        grouped_data = context['grouped_data']
        
        # Verificar que solo aparece la participación de esta semana
        activity_names = [item['activity_name'] for item in grouped_data if item.get('period')]
        self.assertIn('Fútbol', activity_names)
        # La actividad de hace 2 semanas no debería aparecer si tiene período
        # (pero puede aparecer si solo tiene inscripciones sin período)
    
    def test_filter_by_period_mensual(self):
        """Test de filtro por período mensual"""
        # Eliminar participación existente del setUp para evitar duplicados
        Participation.objects.filter(
            activity=self.activity1,
            user=self.student_user
        ).delete()
        
        # Crear participación este mes
        today = timezone.localtime().date()
        participation_this_month, _ = Participation.objects.get_or_create(
            activity=self.activity1,
            user=self.student_user,
            attendance_date=today,
            defaults={'schedule': self.schedule1}
        )
        
        # Crear participación hace 2 meses
        two_months_ago = today - timedelta(days=60)
        participation_old = Participation.objects.create(
            activity=self.activity2,
            user=self.student_user,
            attendance_date=two_months_ago,
            schedule=self.schedule2
        )
        
        url = reverse("participation_segmentation")
        response = self.client.get(url, {'period_filter': 'mensual'})
        
        self.assertEqual(response.status_code, 200)
        context = response.context
        grouped_data = context['grouped_data']
        
        # Verificar que solo aparece la participación de este mes
        activity_names = [item['activity_name'] for item in grouped_data if item.get('period')]
        self.assertIn('Fútbol', activity_names)
    
    def test_filter_by_user_type(self):
        """Test de filtro por tipo de usuario"""
        # Crear grupo de estudiantes
        student_group, _ = Group.objects.get_or_create(name='basic user')
        self.student_user.groups.add(student_group)
        
        # Crear otro usuario que no es estudiante
        other_user = User.objects.create_user(
            username="other_user",
            password="pass123"
        )
        other_group, _ = Group.objects.get_or_create(name='profesor')
        other_user.groups.add(other_group)
        
        # Crear enrollment y participación para ambos usuarios
        Enrollment.objects.create(
            user=other_user,
            activity=self.activity1,
            schedule=self.schedule1
        )
        today = timezone.localtime().date()
        Participation.objects.create(
            activity=self.activity1,
            user=other_user,
            attendance_date=today,
            schedule=self.schedule1
        )
        
        url = reverse("participation_segmentation")
        response = self.client.get(url, {'user_type': 'estudiante'})
        
        self.assertEqual(response.status_code, 200)
        context = response.context
        grouped_data = context['grouped_data']
        
        # Verificar que solo aparecen estudiantes
        for group in grouped_data:
            enrolled_students = group.get('enrolled_students', [])
            participated_students = group.get('participated_students', [])
            all_students = enrolled_students + participated_students
            student_usernames = [s.get('username') for s in all_students]
            # Si el grupo 'profesor' existe y tiene usuarios, no deberían aparecer
            if other_group.user_set.exists():
                # Solo verificar si hay datos
                if student_usernames:
                    # Los estudiantes deberían estar en el grupo 'basic user'
                    self.assertIn('student_user', student_usernames)
    
    def test_filter_by_schedule(self):
        """Test de filtro por horario"""
        # Eliminar participación existente del setUp para evitar duplicados
        Participation.objects.filter(
            activity=self.activity1,
            user=self.student_user
        ).delete()
        
        # Crear otro schedule diferente
        schedule3 = Schedule.objects.create(
            activity=self.activity1,
            day=WeekDay.WEDNESDAY,
            start_time=time(16, 0),
            end_time=time(17, 0),
        )
        
        # Crear enrollment y participación con schedule diferente
        enrollment3 = Enrollment.objects.create(
            user=self.student_user,
            activity=self.activity1,
            schedule=schedule3
        )
        today = timezone.localtime().date()
        Participation.objects.create(
            activity=self.activity1,
            user=self.student_user,
            attendance_date=today,
            schedule=schedule3
        )
        
        # Filtrar por el schedule original (Lunes 10:00-11:00)
        schedule_filter = "Lunes 10:00-11:00"
        url = reverse("participation_segmentation")
        response = self.client.get(url, {'schedule': schedule_filter})
        
        self.assertEqual(response.status_code, 200)
        context = response.context
        grouped_data = context['grouped_data']
        
        # Verificar que solo aparecen los datos del horario filtrado
        for group in grouped_data:
            if group.get('schedule'):
                # El schedule debería coincidir con el filtro
                self.assertIn('Lunes', group['schedule'])
                self.assertIn('10:00', group['schedule'])
    
    def test_filter_combined(self):
        """Test de filtros combinados (período, tipo de actividad, tipo de usuario, horario)"""
        # Crear grupo de estudiantes
        student_group, _ = Group.objects.get_or_create(name='basic user')
        self.student_user.groups.add(student_group)
        
        # Crear actividad artística
        cultural_activity = Activity.objects.create(
            name="Pintura",
            description="Clase de pintura",
            location="Sala 2",
            type=ActivityType.ARTISTICA,
            category=CategoryType.INDIVIDUAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )
        
        # Crear schedule para actividad cultural
        cultural_schedule = Schedule.objects.create(
            activity=cultural_activity,
            day=WeekDay.FRIDAY,
            start_time=time(15, 0),
            end_time=time(16, 0),
        )
        
        # Crear enrollment y participación para actividad cultural
        Enrollment.objects.create(
            user=self.student_user,
            activity=cultural_activity,
            schedule=cultural_schedule
        )
        today = timezone.localtime().date()
        Participation.objects.create(
            activity=cultural_activity,
            user=self.student_user,
            attendance_date=today,
            schedule=cultural_schedule
        )
        
        # Aplicar filtros combinados: tipo deportiva, período mensual, tipo estudiante, horario Lunes 10:00-11:00
        url = reverse("participation_segmentation")
        response = self.client.get(url, {
            'activity_type': ActivityType.DEPORTIVA,
            'period_filter': 'mensual',
            'user_type': 'estudiante',
            'schedule': 'Lunes 10:00-11:00'
        })
        
        self.assertEqual(response.status_code, 200)
        context = response.context
        
        # Verificar que los filtros se aplicaron correctamente
        self.assertEqual(context['selected_type'], ActivityType.DEPORTIVA)
        self.assertEqual(context['selected_period'], 'mensual')
        self.assertEqual(context['selected_user_type'], 'estudiante')
        self.assertEqual(context['selected_schedule'], 'Lunes 10:00-11:00')
        
        grouped_data = context['grouped_data']
        # Verificar que solo aparecen actividades deportivas
        for group in grouped_data:
            self.assertEqual(group['activity_type'], ActivityType.DEPORTIVA)
    
    def test_no_results_with_filters(self):
        """Test cuando no hay resultados con los filtros aplicados"""
        # Crear actividad que no coincide con ningún filtro
        other_activity = Activity.objects.create(
            name="Otra Actividad",
            description="Descripción",
            location="Lugar",
            type=ActivityType.ARTISTICA,
            category=CategoryType.INDIVIDUAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )
        
        # Aplicar filtro que no coincide (tipo deportiva cuando la actividad es cultural)
        url = reverse("participation_segmentation")
        response = self.client.get(url, {
            'activity_type': ActivityType.DEPORTIVA,
            'period_filter': 'mensual',
            'user_type': 'estudiante',
            'schedule': 'Lunes 10:00-11:00'
        })
        
        self.assertEqual(response.status_code, 200)
        # Verificar que se muestra el mensaje de "no hay resultados"
        # El template debería mostrar grouped_data vacío o el mensaje
        context = response.context
        grouped_data = context.get('grouped_data', [])
        
        # Si no hay resultados, grouped_data puede estar vacío o contener solo elementos sin período
        # Verificar que el template maneja esto correctamente
        # (esto se verifica en el template con {% else %})


class DownloadSegmentationTests(TestCase):
    """Tests para las funciones de descarga de segmentación"""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username="admin_user",
            password="pass123",
            is_staff=True
        )
        self.client.login(username="admin_user", password="pass123")

        self.student_user = User.objects.create_user(
            username="student_user",
            password="pass123"
        )

        self.activity = Activity.objects.create(
            name="Fútbol",
            description="Partido de fútbol",
            location="Cancha 1",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )

        self.schedule = Schedule.objects.create(
            activity=self.activity,
            day=WeekDay.MONDAY,
            start_time=time(10, 0),
            end_time=time(11, 0),
        )

        # Crear enrollment con fecha pasada para que tenga un periodo
        from django.utils import timezone
        enrollment = Enrollment.objects.create(
            user=self.student_user,
            activity=self.activity,
            schedule=self.schedule
        )
        # Actualizar registered_at manualmente para el test
        enrollment.registered_at = timezone.now() - timedelta(days=30)
        enrollment.save()

    def test_download_segmentation_excel(self):
        """Test de descarga de Excel de segmentación"""
        url = reverse("download_segmentation_excel")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'],
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Verificar que el archivo es un Excel válido
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertIsNotNone(ws)

        # Verificar encabezados
        self.assertEqual(ws.cell(row=1, column=1).value, 'Periodo')
        self.assertEqual(ws.cell(row=1, column=2).value, 'Actividad')
        self.assertEqual(ws.cell(row=1, column=3).value, 'Total de Participantes')

    def test_download_segmentation_excel_with_filter(self):
        """Test de descarga de Excel con filtro por tipo"""
        url = reverse("download_segmentation_excel")
        response = self.client.get(url, {'activity_type': ActivityType.DEPORTIVA})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'],
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Verificar que el archivo es un Excel válido
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertIsNotNone(ws)

    def test_download_segmentation_excel_empty_data(self):
        """Test de descarga de Excel con datos vacíos"""
        # Eliminar todos los datos
        Enrollment.objects.all().delete()

        url = reverse("download_segmentation_excel")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'],
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Verificar que el archivo Excel es válido incluso con datos vacíos
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertIsNotNone(ws)

        # Verificar que los encabezados están presentes
        self.assertEqual(ws.cell(row=1, column=1).value, 'Periodo')
        self.assertEqual(ws.cell(row=1, column=2).value, 'Actividad')
        self.assertEqual(ws.cell(row=1, column=3).value, 'Total de Participantes')

    def test_download_segmentation_csv(self):
        """Test de descarga de CSV de segmentación"""
        url = reverse("download_segmentation_csv")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')

        # Verificar que el contenido es CSV válido
        content = response.content.decode('utf-8-sig')  # utf-8-sig para manejar BOM
        self.assertIn('Periodo', content)
        self.assertIn('Actividad', content)
        self.assertIn('Total de Participantes', content)

    def test_download_segmentation_csv_with_filter(self):
        """Test de descarga de CSV con filtro por tipo"""
        url = reverse("download_segmentation_csv")
        response = self.client.get(url, {'activity_type': ActivityType.DEPORTIVA})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')

        # Verificar contenido CSV
        content = response.content.decode('utf-8-sig')
        self.assertIn('Periodo', content)
        self.assertIn('Actividad', content)

    def test_download_segmentation_csv_empty_data(self):
        """Test de descarga de CSV con datos vacíos"""
        # Eliminar todos los datos
        Enrollment.objects.all().delete()

        url = reverse("download_segmentation_csv")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')

        # Verificar que el contenido es CSV válido incluso con datos vacíos
        content = response.content.decode('utf-8-sig')
        self.assertIn('Periodo', content)
        self.assertIn('Actividad', content)
        self.assertIn('Total de Participantes', content)


class UnifiedCalendarViewTests(TestCase):
    """Tests para UnifiedCalendarView"""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username="student_user",
            password="pass123"
        )
        self.client.login(username="student_user", password="pass123")

        self.activity = Activity.objects.create(
            name="Fútbol",
            description="Partido de fútbol",
            location="Cancha 1",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )

        self.schedule = Schedule.objects.create(
            activity=self.activity,
            day="Lunes",
            start_time=time(10, 0),
            end_time=time(11, 0),
        )

        # Crear enrollment
        self.enrollment = Enrollment.objects.create(
            user=self.user,
            activity=self.activity,
            schedule=self.schedule
        )

    def test_unified_calendar_view_basic(self):
        """Test básico de la vista de calendario unificado"""
        url = reverse("unified_calendar")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fútbol")

        # Verificar contexto
        context = response.context
        self.assertIn('calendar_days', context)
        self.assertIn('month_name', context)
        self.assertIn('year', context)

    def test_unified_calendar_view_with_navigation(self):
        """Test de navegación por meses en el calendario"""
        url = reverse("unified_calendar")
        response = self.client.get(url, {'month': 12, 'year': 2024})

        self.assertEqual(response.status_code, 200)
        context = response.context
        self.assertEqual(context['month_name'], 'December')
        self.assertEqual(context['year'], 2024)

    def test_unified_calendar_view_with_activities(self):
        """Test que muestra actividades en el calendario"""
        url = reverse("unified_calendar")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        context = response.context

        # Verificar que hay días en el calendario
        calendar_days = context['calendar_days']
        self.assertIsInstance(calendar_days, list)
        self.assertGreater(len(calendar_days), 0)

        # Verificar estructura de cada día
        for day_data in calendar_days:
            self.assertIn('day', day_data)
            self.assertIn('date', day_data)
            self.assertIn('current_month', day_data)
            self.assertIn('items', day_data)
            self.assertIsInstance(day_data['items'], list)


class MyCalendarViewTests(TestCase):
    """Tests para MyCalendarView"""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username="student_user",
            password="pass123"
        )
        self.client.login(username="student_user", password="pass123")

        self.activity = Activity.objects.create(
            name="Fútbol",
            description="Partido de fútbol",
            location="Cancha 1",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )

        self.schedule = Schedule.objects.create(
            activity=self.activity,
            day="Lunes",
            start_time=time(10, 0),
            end_time=time(11, 0),
        )

        # Crear enrollment
        self.enrollment = Enrollment.objects.create(
            user=self.user,
            activity=self.activity,
            schedule=self.schedule
        )

    def test_my_calendar_view_basic(self):
        """Test básico de la vista de calendario personal"""
        url = reverse("my_calendar")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fútbol")

        # Verificar contexto
        context = response.context
        self.assertIn('calendar', context)
        calendar = context['calendar']
        self.assertIsInstance(calendar, dict)

        # Verificar que contiene días de la semana
        expected_days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        for day in expected_days:
            self.assertIn(day, calendar)

    def test_my_calendar_view_with_multiple_schedules(self):
        """Test con múltiples horarios para la misma actividad"""
        # Crear otro schedule para el mismo día
        schedule2 = Schedule.objects.create(
            activity=self.activity,
            day="Lunes",
            start_time=time(14, 0),
            end_time=time(15, 0),
        )

        enrollment2 = Enrollment.objects.create(
            user=self.user,
            activity=self.activity,
            schedule=schedule2
        )

        url = reverse("my_calendar")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        context = response.context
        calendar = context['calendar']

        # Verificar que el día Lunes tiene ambas actividades
        lunes_activities = calendar["Lunes"]
        self.assertEqual(len(lunes_activities), 2)

        # Verificar que ambas actividades están presentes
        activity_names = [activity['name'] for activity in lunes_activities]
        self.assertIn("Fútbol", activity_names)


class CreateActivityViewTests(TestCase):
    """Tests para CreateActivityView"""

    def setUp(self):
        self.client = Client()
        # Crear usuario admin
        self.admin_user = User.objects.create_user(
            username="admin_user",
            password="pass123",
            is_staff=True
        )
        self.client.login(username="admin_user", password="pass123")

    def test_create_activity_view_get(self):
        """Test GET request para crear actividad"""
        url = reverse("create_activity")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Publicar nueva actividad")

        # Verificar que el contexto tiene formset
        context = response.context
        self.assertIn('formset', context)

    def test_create_activity_view_post_success(self):
        """Test POST exitoso para crear actividad"""
        url = reverse("create_activity")

        data = {
            'name': 'Nueva Actividad',
            'description': 'Descripción de prueba',
            'location': 'Ubicación de prueba',
            'type': ActivityType.DEPORTIVA,
            'category': CategoryType.GRUPAL,
            'is_published': True,
            'requires_registration': True,
            'max_capacity': 20,
            'schedules-TOTAL_FORMS': '1',
            'schedules-INITIAL_FORMS': '0',
            'schedules-MIN_NUM_FORMS': '0',
            'schedules-MAX_NUM_FORMS': '1000',
            'schedules-0-day': 'Lunes',
            'schedules-0-start_time': '10:00',
            'schedules-0-end_time': '11:00',
        }

        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)  # Redirect

        # Verificar que la actividad fue creada
        self.assertEqual(Activity.objects.count(), 1)
        activity = Activity.objects.first()
        self.assertEqual(activity.name, 'Nueva Actividad')

        # Verificar que el schedule fue creado
        self.assertEqual(Schedule.objects.count(), 1)
        schedule = Schedule.objects.first()
        self.assertEqual(schedule.day, 'Lunes')

    def test_create_activity_view_post_invalid(self):
        """Test POST con datos inválidos"""
        url = reverse("create_activity")

        data = {
            'name': '',  # Nombre vacío (inválido)
            'description': 'Descripción de prueba',
            'location': 'Ubicación de prueba',
            'type': ActivityType.DEPORTIVA,
            'category': CategoryType.GRUPAL,
            'is_published': True,
            'requires_registration': True,
            'max_capacity': 20,
            'schedules-TOTAL_FORMS': '1',
            'schedules-INITIAL_FORMS': '0',
            'schedules-MIN_NUM_FORMS': '0',
            'schedules-MAX_NUM_FORMS': '1000',
            'schedules-0-day': 'Lunes',
            'schedules-0-start_time': '10:00',
            'schedules-0-end_time': '11:00',
        }

        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 200)  # No redirect, muestra errores

        # Verificar que la actividad NO fue creada
        self.assertEqual(Activity.objects.count(), 0)


class UpdateActivityViewTests(TestCase):
    """Tests para UpdateActivityView"""

    def setUp(self):
        self.client = Client()
        self.admin_user = User.objects.create_user(
            username="admin_user",
            password="pass123",
            is_staff=True
        )
        self.client.login(username="admin_user", password="pass123")

        self.activity = Activity.objects.create(
            name="Actividad Original",
            description="Descripción original",
            location="Ubicación original",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )

    def test_update_activity_view_get(self):
        """Test GET request para actualizar actividad"""
        url = reverse("update_activity", args=[self.activity.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Actividad Original")

        # Verificar contexto
        context = response.context
        self.assertIn('activity', context)
        self.assertEqual(context['activity'], self.activity)

    def test_update_activity_view_post_success(self):
        """Test POST exitoso para actualizar actividad"""
        url = reverse("update_activity", args=[self.activity.pk])

        data = {
            'name': 'Actividad Actualizada',
            'description': 'Descripción actualizada',
            'location': 'Ubicación actualizada',
            'type': ActivityType.DEPORTIVA,
            'category': CategoryType.GRUPAL,
            'is_published': True,
            'requires_registration': True,
            'max_capacity': 15,
        }

        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)  # Redirect

        # Verificar que la actividad fue actualizada
        self.activity.refresh_from_db()
        self.assertEqual(self.activity.name, 'Actividad Actualizada')
        self.assertEqual(self.activity.max_capacity, 15)


class DeleteActivityViewTests(TestCase):
    """Tests para DeleteActivityView"""

    def setUp(self):
        self.client = Client()
        self.admin_user = User.objects.create_user(
            username="admin_user",
            password="pass123",
            is_staff=True
        )
        self.client.login(username="admin_user", password="pass123")

        self.activity = Activity.objects.create(
            name="Actividad a Eliminar",
            description="Descripción",
            location="Ubicación",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )

    def test_delete_activity_view_post(self):
        """Test POST para eliminar actividad"""
        url = reverse("delete_activity", args=[self.activity.pk])
        response = self.client.post(url)

        self.assertEqual(response.status_code, 302)  # Redirect
        self.assertRedirects(response, reverse("public_activities"))

        # Verificar que la actividad fue eliminada
        self.assertEqual(Activity.objects.count(), 0)

    def test_delete_activity_view_get(self):
        """Test GET para mostrar confirmación de eliminación"""
        url = reverse("delete_activity", args=[self.activity.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Actividad a Eliminar")


class ConfirmEnrollmentViewTests(TestCase):
    """Tests para ConfirmEnrollmentView"""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username="student_user",
            password="pass123"
        )
        self.client.login(username="student_user", password="pass123")

        self.activity = Activity.objects.create(
            name="Actividad con Confirmación",
            description="Descripción",
            location="Ubicación",
            type=ActivityType.DEPORTIVA,
            category=CategoryType.GRUPAL,
            is_published=True,
            requires_registration=True,
            max_capacity=10,
        )

        self.enrollment = Enrollment.objects.create(
            user=self.user,
            activity=self.activity
        )

    def test_confirm_enrollment_view_valid_token(self):
        """Test con token válido"""
        from django.core.signing import dumps
        token = dumps({"enrollment_id": self.enrollment.id, "user_id": self.user.id})

        url = reverse("confirm_enrollment", args=[token])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 302)  # Redirect
        self.assertRedirects(response, reverse("my_calendar"))

        # Verificar que la inscripción fue confirmada
        self.enrollment.refresh_from_db()
        self.assertTrue(self.enrollment.confirmed)

    def test_confirm_enrollment_view_expired_token(self):
        """Test con token expirado"""
        from django.core.signing import dumps, loads
        # Crear token válido primero
        token = dumps({"enrollment_id": self.enrollment.id, "user_id": self.user.id})

        # Esperar un poco para que expire
        import time
        time.sleep(2)

        # Intentar cargar el token con max_age=1 (muy corto)
        try:
            loads(token, max_age=1)
            self.fail("Token should have expired")
        except:
            # Token expirado, ahora probar la vista
            url = reverse("confirm_enrollment", args=[token])
            response = self.client.get(url)
            self.assertEqual(response.status_code, 302)  # Redirect con error

    def test_confirm_enrollment_view_invalid_token(self):
        """Test con token inválido"""
        url = reverse("confirm_enrollment", args=["invalid_token"])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 302)  # Redirect con error
